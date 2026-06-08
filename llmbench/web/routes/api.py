"""REST API routes."""
import logging
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query, BackgroundTasks
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
import io
import json
import asyncio
import re
import random
import time
from urllib.parse import quote

logger = logging.getLogger(__name__)

from ..database import get_db, AsyncSessionLocal
from ..crud import BenchmarkCRUD, ScenarioCRUD, ServerCRUD, ChallengeCRUD
from ..schemas import (
    BenchmarkCreate,
    BenchmarkResponse,
    BenchmarkListItem,
    BenchmarkStatus,
    ServerCreate,
    ServerResponse,
    ChallengeResponse,
    ChallengeListItem,
)
from ..models import Benchmark, Challenge
from ..tasks import run_benchmark_task

router = APIRouter()

def _parse_path_list(raw: str) -> list[str]:
    """Parse a user-supplied list of file paths.

    Handles comma-separated, newline-separated, CSV-quoted values,
    and adjacent quoted paths without separators
    (e.g. "C:\\path1","C:\\path2" or "C:\\path1""C:\\path2").
    """
    import csv as _csv
    # Use regex to extract all quoted or unquoted path tokens
    # Matches: "any chars" or non-comma/newline sequences
    tokens = re.findall(r'"([^"]+)"|\'([^\']+)\'|([^,\n\r"\']+)', raw)
    paths: list[str] = []
    for groups in tokens:
        val = (groups[0] or groups[1] or groups[2] or "").strip()
        if val:
            paths.append(val)
    return paths or None


_DEFAULT_ASPECT_LABELS = ["人文", "歷史", "政治", "社會", "國際", "科技"]
_DEFAULT_SCHOOL_SUBJECTS = ["國文", "英語", "數學", "自然", "社會", "理化", "生物", "歷史", "地理", "公民"]
_ASPECT_KEYWORDS = {
    "人文": [
        "人文", "文化", "文學", "哲學", "藝術", "宗教", "語言", "博物館", "展覽", "戲劇", "歌仔戲", "表演藝術",
        # 社會科：族群文化、地理人文
        "族群", "原住民", "客家", "閩南", "眷村", "民俗", "祭典", "節慶", "傳統", "風俗", "習俗",
        "地理", "地形", "氣候", "河川", "山脈", "平原", "盆地", "海岸", "地圖", "經緯度",
        "聚落", "都市", "農村", "建築", "景觀",
    ],
    "歷史": [
        "歷史", "古代", "近代", "戰役", "王朝", "年代", "考古", "史料", "史學", "年表",
        "清朝", "日治", "日據", "戰後", "光復", "民國", "清廷", "鄭氏", "荷蘭", "西班牙",
        "朝代", "皇帝", "條約", "開港", "割讓", "統治", "殖民", "抗日", "二二八", "戒嚴",
        "解嚴", "改革", "革命", "起義", "遷台", "建國", "開疆", "拓墾", "屯墾",
        "史前", "考古", "出土", "文物", "遺址", "石器", "陶器",
        # 學測常見：中國史、世界史
        "世紀", "帝國", "封建", "工業革命", "文藝復興", "啟蒙", "冷戰", "二戰", "一戰",
        "大航海", "貿易路線", "殖民地", "獨立運動",
        "漢朝", "唐朝", "宋朝", "明朝", "清代", "隋唐", "秦漢",
        "以前", "時期", "時代", "當時", "那時",
    ],
    "政治": [
        "政治", "政府", "總統", "立法院", "議會", "選舉", "政黨", "政策", "部會", "公投",
        "國會", "內閣", "閣揆", "行政院", "縣市長", "外交政策", "朝野",
        # 社會科：公民、政府體制
        "憲法", "法律", "法規", "人權", "民主", "自由", "平等", "公民", "權利", "義務",
        "三權", "五權", "行政", "立法", "司法", "考試院", "監察院",
        "地方自治", "中央", "縣市", "鄉鎮", "議員", "里長",
        "聯合國", "主權", "國際法", "外交", "邦交",
    ],
    "社會": [
        "社會", "民生", "醫療", "勞工", "居住", "家庭", "校園", "治安", "司法", "福利",
        "社福", "長照", "住宅", "房價", "交通", "公安", "食安", "就業", "教育現場",
        "健保", "勞保", "年金", "保險", "醫院", "診所", "疫情", "防疫", "托育", "育兒",
        "老人", "長者", "弱勢", "低收入", "扶助", "補助", "少子化", "高齡化",
        "人口", "移民", "生育率", "失業", "工時", "薪資", "所得",
        "衛生", "全民健保", "健康保險", "社會保險", "社會安全",
        # 社會科：經濟、產業、生活
        "經濟", "產業", "農業", "工業", "商業", "服務業", "貿易", "出口", "進口",
        "消費", "生產", "市場", "物價", "貧富",
        "環境", "汙染", "生態", "永續", "資源", "能源", "碳排", "再生能源",
        "教育", "學校", "課程", "升學", "師生", "義務教育",
    ],
    "國際": [
        "國際", "全球", "外交", "聯合國", "美國", "中國", "日本", "歐盟", "俄羅斯", "烏克蘭",
        "兩岸", "中華人民共和國", "中共", "北京", "印太", "台灣海峽", "南海",
        "邦交國", "建交", "斷交", "軍售", "制裁",
        # 社會科：世界地理、國際關係
        "亞洲", "歐洲", "非洲", "美洲", "大洋洲", "中東", "東南亞", "東北亞",
        "世界", "各國", "跨國", "全球化", "區域", "同盟", "組織",
        "氣候變遷", "全球暖化", "溫室效應", "永續發展",
        "東協", "nato", "apec", "wto",
    ],
    "科技": [
        "科技", "技術", "ai", "人工智慧", "晶片", "半導體", "軟體", "網路", "手機", "電腦",
        "台積電", "聯發科", "鴻海", "宏碁", "華碩", "緯創", "中華電信",
        "晶圓", "積體電路", "製程", "奈米", "晶圓代工", "封裝測試",
        "5g", "電動車", "太陽能", "光電", "電池", "充電",
        # 社會科：科技與產業
        "資訊", "數位", "網際", "通訊", "電信", "衛星",
        "交通建設", "鐵路", "高鐵", "捷運", "港口", "機場",
        "科學園區", "工研院", "研發", "專利", "新創",
    ],
}


def _split_threads_sections(text: str) -> tuple[str, str]:
    """Split combined Threads content into post text and reply text."""
    raw = (text or "").strip()
    if not raw:
        return "", ""

    if "【主題】" not in raw:
        return raw, ""

    body = raw.split("【主題】", 1)[1].strip()
    match = re.search(r"\n\s*【留言[^】]*】\n?", body)
    if not match:
        return body.strip(), ""

    post_text = body[:match.start()].strip()
    replies_text = body[match.end():].strip()
    return post_text, replies_text

def _is_supported_stance_language(text: str) -> bool:
    """Return True when the text is primarily Chinese or English."""
    raw = (text or "").strip()
    if not raw:
        return False

    zh_count = len(re.findall(r"[\u4e00-\u9fff]", raw))
    en_count = len(re.findall(r"[A-Za-z]", raw))
    arabic_count = len(re.findall(r"[\u0600-\u06FF]", raw))

    meaningful = zh_count + en_count + arabic_count
    if meaningful == 0:
        return False
    if arabic_count > max(zh_count, en_count):
        return False
    return (zh_count + en_count) >= meaningful * 0.6


def _is_rejectable_stance_topic(topic: str, content: str) -> bool:
    topic_norm = (topic or "").strip()
    content_norm = (content or "").strip()
    lowered = topic_norm.lower()
    generic_topics = {
        "prompt",
        "threads",
        "post",
        "tweet",
        "ai",
        "chatgpt",
        "gemini",
    }
    prompt_markers = [
        "a high-resolution",
        "cinematic",
        "saved face reference",
        "style & attire",
        "black canvas utility jacket",
        "voluminous dark espresso hair",
        "sharp, angular jawline",
        "prompt",
        "midjourney",
        "stable diffusion",
    ]

    if len(topic_norm) < 8:
        return True
    if lowered in generic_topics:
        return True
    if any(marker in lowered for marker in prompt_markers):
        return True
    if content_norm and lowered == content_norm.lower():
        return True
    if any(marker in content_norm.lower() for marker in prompt_markers) and ("描述" in topic_norm or "prompt" in lowered):
        return True
    return False


def _parse_selected_aspects(raw: Optional[str]) -> list[str]:
    if not raw or not raw.strip():
        return list(_DEFAULT_ASPECT_LABELS)

    selected: list[str] = []
    seen: set[str] = set()
    for part in re.split(r"[\n,]+", raw):
        label = part.strip()
        if not label or label in seen or label not in _DEFAULT_ASPECT_LABELS:
            continue
        seen.add(label)
        selected.append(label)
    return selected or list(_DEFAULT_ASPECT_LABELS)


_TITLE_ASPECT_MAP: dict[str, str] = {
    # 人文（地理、文化、生活）
    "台北市": "人文", "高雄市": "人文", "台中市": "人文", "台南市": "人文",
    "新北市": "人文", "桃園市": "人文", "基隆市": "人文", "宜蘭縣": "人文",
    "花蓮縣": "人文", "台東縣": "人文", "澎湖縣": "人文", "金門縣": "人文",
    "馬祖列島": "人文", "玉山": "人文", "阿里山": "人文", "日月潭": "人文",
    "太魯閣": "人文", "墾丁": "人文", "台灣海峽": "人文", "台北101": "人文",
    "夜市": "人文", "便利商店": "人文", "台灣料理": "人文", "珍珠奶茶": "人文",
    "台灣棒球": "人文", "台灣電影": "人文", "台灣流行音樂": "人文", "台灣文學": "人文",
    "九份": "人文", "台灣鐵路": "人文", "原住民族傳統": "人文", "布袋戲": "人文",
    "歌仔戲": "人文", "媽祖": "人文", "城隍": "人文", "台灣廟宇": "人文",
    "傳統節日": "人文", "農曆新年": "人文", "元宵節": "人文", "端午節": "人文",
    "中元節": "人文", "中秋節": "人文", "清明節": "人文", "台灣民間信仰": "人文",
    "台灣宗教": "人文", "廟會": "人文", "豐年祭": "人文",
    "臺灣原住民族": "人文", "原住民族": "人文", "便利商店": "人文",
    # 社會
    "全民健康保險": "社會", "健康保險": "社會", "勞工保險": "社會", "勞工退休": "社會",
    "長期照顧": "社會", "長期照護": "社會", "社會福利": "社會", "少子化": "社會", "老年化": "社會",
    "高齡化": "社會", "生育率": "社會", "失業率": "社會", "失業保險": "社會", "住宅政策": "社會",
    "房價": "社會", "食品安全": "社會", "能源政策": "社會", "碳排放": "社會",
    "勞動市場": "社會", "勞動基準法": "社會", "勞動部": "社會",
    "台灣人口": "社會", "全民教育": "社會", "台灣交通": "社會", "義務教育": "社會",
    "中華民國刑法": "社會", "中華民國民法": "社會", "性別平等": "社會", "同性婚姻": "社會",
    "國民年金": "社會", "衛生福利部": "社會", "教育部": "社會", "內政部": "社會",
    "國民身分證": "社會", "身心障礙": "社會", "國立臺灣大學": "社會", "臺灣大學": "社會",
    "台灣教育": "社會",
    # 科技
    "台積電": "科技", "積體電路製造": "科技", "聯華電子": "科技", "聯發科": "科技",
    "鴻海": "科技", "半導體": "科技", "科學工業園區": "科技", "工業技術研究院": "科技",
    "台灣電力": "科技", "高速鐵路": "科技", "台北捷運": "科技", "太陽能": "科技",
    "人工智慧": "科技", "電動車": "科技", "中華電信": "科技",
    "華碩": "科技", "宏碁": "科技", "再生能源": "科技",
    # 政治
    "中華民國總統": "政治", "行政院": "政治", "立法院": "政治", "憲法": "政治", "增修條文": "政治",
    "國慶日": "政治", "司法院": "政治", "考試院": "政治", "監察院": "政治",
    "國旗": "政治", "國徽": "政治", "總統府": "政治", "外交": "政治",
    # 國際
    "中華人民共和國": "國際", "北京市": "國際", "上海市": "國際", "習近平": "國際", "人民幣": "國際",
    "普通話": "國際", "全國人民代表大會": "國際", "國務院": "國際", "中國共產黨": "國際",
    "中華人民共和國國旗": "國際", "中華人民共和國國徽": "國際", "中華人民共和國國慶日": "國際",
    "日本": "國際", "東京都": "國際", "日圓": "國際", "日本國憲法": "國際", "日本天皇": "國際",
    "美國": "國際", "美元": "國際", "華盛頓": "國際", "美國總統": "國際", "美國國會": "國際",
    "大韓民國": "國際", "首爾": "國際",
}


def _infer_aspect_from_text(*parts: Optional[str]) -> str:
    # First part is typically the article title — use title map for strong signal
    title = (parts[0] or "").strip() if parts else ""
    for pattern, aspect in _TITLE_ASPECT_MAP.items():
        if pattern in title:
            return aspect

    text = " ".join((part or "").strip() for part in parts if part).lower()
    if not text:
        return "社會"

    best_label = "社會"
    best_score = 0
    for label in _DEFAULT_ASPECT_LABELS:
        score = 0
        for keyword in _ASPECT_KEYWORDS.get(label, []):
            key = keyword.lower()
            if key in text:
                score += max(1, text.count(key))
        if score > best_score:
            best_label = label
            best_score = score
    return best_label


def _filter_jsonl_by_aspects(data_jsonl: str, selected_aspects: list[str]) -> tuple[str, int, int]:
    lines_kept: list[str] = []
    total_rows = 0

    for raw_line in data_jsonl.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        total_rows += 1
        try:
            row = json.loads(line)
        except Exception:
            aspect = _infer_aspect_from_text(line)
        else:
            aspect = _infer_aspect_from_text(
                str(row.get("title") or ""),
                str(row.get("text") or ""),
                str(row.get("content") or ""),
                str(row.get("keyword") or ""),
                str(row.get("source_category") or ""),
            )
        if aspect in selected_aspects:
            lines_kept.append(line)

    return "\n".join(lines_kept), total_rows, len(lines_kept)


def _sample_jsonl_rows(
    data_jsonl: str,
    selected_aspects: list[str],
    sample_size: int,
    sample_scope: str,
) -> tuple[str, int]:
    lines = [line.strip() for line in data_jsonl.splitlines() if line.strip()]
    if not lines:
        return "", 0

    buckets: dict[str, list[str]] = {label: [] for label in selected_aspects}
    fallbacks: list[str] = []
    for line in lines:
        try:
            row = json.loads(line)
        except Exception:
            aspect = _infer_aspect_from_text(line)
        else:
            aspect = _infer_aspect_from_text(
                str(row.get("title") or ""),
                str(row.get("text") or ""),
                str(row.get("content") or ""),
                str(row.get("keyword") or ""),
                str(row.get("source_category") or ""),
            )
        if aspect in buckets:
            buckets[aspect].append(line)
        else:
            fallbacks.append(line)

    bounded_size = max(1, min(sample_size, 100))
    if sample_scope != "per_aspect":
        guaranteed: list[str] = []
        seen: set[str] = set()
        remaining_pool: list[str] = []

        for label in selected_aspects:
            choices = buckets.get(label, [])
            if not choices:
                continue
            line = random.choice(choices)
            if line in seen:
                continue
            seen.add(line)
            guaranteed.append(line)

        for group in list(buckets.values()) + [fallbacks]:
            for line in group:
                if line not in seen:
                    remaining_pool.append(line)

        if len(guaranteed) >= bounded_size:
            chosen = guaranteed[:bounded_size]
        else:
            extra = random.sample(remaining_pool, k=min(len(remaining_pool), bounded_size - len(guaranteed)))
            chosen = guaranteed + extra
        return "\n".join(chosen), len(chosen)

    sampled: list[str] = []
    seen: set[str] = set()
    base_quota = bounded_size // max(1, len(selected_aspects))
    remainder = bounded_size % max(1, len(selected_aspects))
    remaining_pool: list[str] = []

    for index, label in enumerate(selected_aspects):
        choices = list(buckets.get(label, []))
        quota = base_quota + (1 if index < remainder else 0)
        take = min(len(choices), quota)
        for line in random.sample(choices, k=take):
            if line in seen:
                continue
            seen.add(line)
            sampled.append(line)
        for line in choices:
            if line not in seen:
                remaining_pool.append(line)

    for line in fallbacks:
        if line not in seen:
            remaining_pool.append(line)

    if len(sampled) < bounded_size and remaining_pool:
        extra = random.sample(remaining_pool, k=min(len(remaining_pool), bounded_size - len(sampled)))
        sampled.extend(extra)

    if not sampled and fallbacks:
        sampled = random.sample(fallbacks, k=min(len(fallbacks), bounded_size))
    return "\n".join(sampled), len(sampled)


def _collect_aspect_breakdown(
    items: list[dict],
    model_names: list[str],
    labels: Optional[list[str]] = None,
    use_keyword: bool = False,
) -> dict[str, dict[str, dict[str, Optional[float]]]]:
    effective_labels = labels if labels else _DEFAULT_ASPECT_LABELS
    fallback_label = effective_labels[0] if effective_labels else "其他"

    per_model_aspect_scores: dict[str, dict[str, list[float]]] = {
        model_name: {label: [] for label in effective_labels}
        for model_name in model_names
    }

    for item in items:
        model_name = str(item.get("model_name") or "")
        if not model_name:
            continue
        if model_name not in per_model_aspect_scores:
            per_model_aspect_scores[model_name] = {label: [] for label in effective_labels}

        if use_keyword:
            aspect = str(item.get("keyword") or item.get("aspect") or "").strip()
        else:
            aspect = str(item.get("aspect") or "").strip() or _infer_aspect_from_text(
                str(item.get("title") or ""),
                str(item.get("topic") or ""),
                str(item.get("content") or ""),
                str(item.get("full_content") or ""),
            )
        if aspect not in effective_labels:
            aspect = fallback_label
        item["aspect"] = aspect
        score = item.get("score")
        if isinstance(score, (int, float)):
            per_model_aspect_scores[model_name][aspect].append(float(score))

    return {
        model_name: {
            "scores": {
                label: (round(sum(values) / len(values), 3) if values else None)
                for label, values in aspect_map.items()
            },
            "counts": {
                label: len(values)
                for label, values in aspect_map.items()
            },
        }
        for model_name, aspect_map in per_model_aspect_scores.items()
    }


def _sort_leaderboard_rows(rows: list[dict]) -> list[dict]:
    sorted_rows = sorted(
        rows,
        key=lambda row: (
            -(row.get("avg_score") if row.get("avg_score") is not None else -1),
            row.get("error_count", 0),
            str(row.get("display_name") or row.get("model_name") or ""),
        ),
    )
    return [
        {
            **row,
            "rank": idx + 1,
        }
        for idx, row in enumerate(sorted_rows)
    ]


def _safe_download_filename(name: str, suffix: str) -> str:
    base = re.sub(r'[<>:"/\\|?*\r\n]+', "_", (name or "").strip())
    base = base.strip(". ")
    if not base:
        base = "challenge"
    return f"{base}{suffix}"


def _build_content_disposition(filename: str) -> str:
    fallback = filename.encode("ascii", "ignore").decode("ascii").strip()
    fallback = re.sub(r"\s+", " ", fallback)
    fallback = fallback or "download"
    return f"attachment; filename=\"{fallback}\"; filename*=UTF-8''{quote(filename)}"


def _diversify_by_article(items: list[dict], n: int) -> list[dict]:
    """Round-robin interleave by article, then de-cluster by title + 4-char statement prefix."""
    by_title: dict[str, list[dict]] = {}
    for item in items:
        key = item.get("title") or "__unknown__"
        by_title.setdefault(key, []).append(item)
    cap = max(1, n // 10)
    buckets = [random.sample(group, k=min(cap, len(group))) for group in by_title.values()]
    random.shuffle(buckets)
    pool: list[dict] = []
    while len(pool) < n:
        added = False
        for bucket in buckets:
            if bucket and len(pool) < n:
                pool.append(bucket.pop(0))
                added = True
        if not added:
            break

    # De-cluster: no two adjacent questions share title OR 4-char statement prefix
    def stmt_prefix(item: dict) -> str:
        ref_raw = item.get("reference_answer", "")
        try:
            ref = json.loads(ref_raw) if isinstance(ref_raw, str) else ref_raw
            stmt = ref.get("statement", "") if isinstance(ref, dict) else ""
        except Exception:
            stmt = ""
        return stmt[:3]

    remaining = list(pool)
    result: list[dict] = []
    while remaining:
        prev = result[-1] if result else None
        prev_title = prev.get("title", "") if prev else ""
        prev_pfx = stmt_prefix(prev) if prev else ""
        # 1st: different title AND different prefix
        ni = next(
            (i for i, it in enumerate(remaining)
             if it.get("title", "") != prev_title and stmt_prefix(it) != prev_pfx),
            None,
        )
        # 2nd: at least different prefix
        if ni is None:
            ni = next((i for i, it in enumerate(remaining) if stmt_prefix(it) != prev_pfx), None)
        # 3rd: at least different title
        if ni is None:
            ni = next((i for i, it in enumerate(remaining) if it.get("title", "") != prev_title), 0)
        result.append(remaining.pop(ni))
    return result


def _compute_aspect_stats(data_jsonl: str) -> dict[str, int]:
    counts = {label: 0 for label in _DEFAULT_ASPECT_LABELS}
    for raw_line in data_jsonl.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        try:
            row = json.loads(line)
        except Exception:
            aspect = _infer_aspect_from_text(line)
        else:
            aspect = _infer_aspect_from_text(
                str(row.get("title") or ""),
                str(row.get("text") or ""),
                str(row.get("content") or ""),
                str(row.get("keyword") or ""),
                str(row.get("source_category") or ""),
            )
        if aspect not in counts:
            aspect = "社會"
        counts[aspect] += 1
    return counts


def _parse_selected_subjects(raw: Optional[str]) -> list[str]:
    if not raw or not raw.strip():
        return list(_DEFAULT_SCHOOL_SUBJECTS)
    selected: list[str] = []
    seen: set[str] = set()
    for part in re.split(r"[\n,]+", raw):
        label = part.strip()
        if not label or label in seen:
            continue
        seen.add(label)
        selected.append(label)
    return selected or list(_DEFAULT_SCHOOL_SUBJECTS)


def _filter_jsonl_by_subjects(data_jsonl: str, selected_subjects: list[str]) -> tuple[str, int, int]:
    """Filter school_qa JSONL rows by keyword (subject) field."""
    lines_kept: list[str] = []
    total_rows = 0
    subject_set = set(selected_subjects)
    for raw_line in data_jsonl.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        total_rows += 1
        try:
            row = json.loads(line)
            keyword = str(row.get("keyword") or "")
        except Exception:
            keyword = ""
        if keyword in subject_set:
            lines_kept.append(line)
    return "\n".join(lines_kept), total_rows, len(lines_kept)


def _compute_subject_stats(data_jsonl: str, selected_subjects: list[str]) -> dict[str, int]:
    """Count school_qa JSONL rows per subject using keyword field."""
    counts = {label: 0 for label in selected_subjects}
    for raw_line in data_jsonl.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        try:
            row = json.loads(line)
            keyword = str(row.get("keyword") or "")
        except Exception:
            keyword = ""
        if keyword in counts:
            counts[keyword] += 1
    return counts


def _sample_jsonl_rows_by_subject(
    data_jsonl: str,
    selected_subjects: list[str],
    sample_size: int,
    sample_scope: str,
) -> tuple[str, int]:
    """Balanced sampling by subject (keyword field) for school_qa challenges."""
    lines = [line.strip() for line in data_jsonl.splitlines() if line.strip()]
    if not lines:
        return "", 0

    buckets: dict[str, list[str]] = {label: [] for label in selected_subjects}
    fallbacks: list[str] = []
    for line in lines:
        try:
            row = json.loads(line)
            keyword = str(row.get("keyword") or "")
        except Exception:
            keyword = ""
        if keyword in buckets:
            buckets[keyword].append(line)
        else:
            fallbacks.append(line)

    bounded_size = max(1, min(sample_size, 100))
    if sample_scope != "per_aspect":
        guaranteed: list[str] = []
        seen: set[str] = set()
        remaining_pool: list[str] = []
        for label in selected_subjects:
            choices = buckets.get(label, [])
            if not choices:
                continue
            line = random.choice(choices)
            if line in seen:
                continue
            seen.add(line)
            guaranteed.append(line)
        for group in list(buckets.values()) + [fallbacks]:
            for line in group:
                if line not in seen:
                    remaining_pool.append(line)
        if len(guaranteed) >= bounded_size:
            chosen = guaranteed[:bounded_size]
        else:
            random.shuffle(remaining_pool)
            chosen = (guaranteed + remaining_pool)[:bounded_size]
    else:
        n_subjects = len(selected_subjects)
        per_subject = max(1, bounded_size // n_subjects) if n_subjects else bounded_size
        chosen = []
        seen_lines: set[str] = set()
        for label in selected_subjects:
            choices = buckets.get(label, [])
            random.shuffle(choices)
            for line in choices[:per_subject]:
                if line not in seen_lines:
                    seen_lines.add(line)
                    chosen.append(line)
        random.shuffle(chosen)
        chosen = chosen[:bounded_size]

    return "\n".join(chosen), len(chosen)


@router.get("/stats")
async def get_dashboard_stats(db: AsyncSession = Depends(get_db)):
    """Get dashboard statistics."""
    total = await BenchmarkCRUD.count_all(db)
    running = await BenchmarkCRUD.count_all(db, status="running")
    completed = await BenchmarkCRUD.count_all(db, status="completed")
    failed = await BenchmarkCRUD.count_all(db, status="failed")
    pending = await BenchmarkCRUD.count_all(db, status="pending")

    # Get recent benchmarks
    recent = await BenchmarkCRUD.list_all(db, limit=5, offset=0)

    return {
        "total_benchmarks": total,
        "running_count": running,
        "completed_count": completed,
        "failed_count": failed,
        "pending_count": pending,
        "recent_benchmarks": [
            {
                "uuid": b.uuid,
                "name": b.name,
                "status": b.status,
                "created_at": b.created_at.isoformat(),
                "runtime_sec": b.runtime_sec,
            }
            for b in recent
        ],
    }


# Simple in-memory cache for /test-runs (TTL 60 seconds).
# Parsing all results_jsonl can take >10 seconds on large DBs, so we cache.
_test_runs_cache: dict = {"data": None, "expires_at": 0.0}


@router.get("/test-runs")
async def list_test_runs(db: AsyncSession = Depends(get_db)):
    """Aggregate challenge test history from results_jsonl into per-run records.

    Groups rows by (challenge_uuid, model_name, hour-bucket) so each "run"
    corresponds to one generate/test session. Result cached 60 seconds.
    """
    import time
    from sqlalchemy import select
    from sqlalchemy.orm import defer
    from collections import defaultdict

    # Serve from cache if still fresh
    now = time.monotonic()
    if _test_runs_cache["data"] is not None and _test_runs_cache["expires_at"] > now:
        return _test_runs_cache["data"]
    from datetime import datetime

    # Only load uuid, name, results_jsonl — defer other heavy text columns
    query = (
        select(Challenge)
        .options(
            defer(Challenge.data_jsonl),
            defer(Challenge.participant_scores_jsonl),
        )
        .order_by(Challenge.updated_at.desc())
    )
    result = await db.execute(query)
    challenges = list(result.scalars().all())

    # Snapshot the small data we need from each ORM object first
    # (uuid/name/task_type/results_jsonl), then do all the heavy parsing in
    # a worker thread so the asyncio event loop stays free for other requests.
    snapshots = [(ch.uuid, ch.name, ch.task_type, ch.results_jsonl or "") for ch in challenges]

    def _parse_all() -> list[dict]:
        out: list[dict] = []
        for ch_uuid, ch_name, ch_task, results_jsonl in snapshots:
            if not results_jsonl:
                continue
            groups: dict[tuple, dict] = defaultdict(
                lambda: {"count": 0, "scores": [], "first_ts": None, "last_ts": None, "errors": 0}
            )
            for line in results_jsonl.splitlines():
                line = line.strip()
                if not line:
                    continue
                try:
                    row = json.loads(line)
                except Exception:
                    continue
                model_name = row.get("model_name") or row.get("model") or "unknown"
                ts_raw = row.get("finished_at") or row.get("requested_at") or ""
                try:
                    ts = datetime.fromisoformat(ts_raw.replace("Z", "+00:00")) if ts_raw else None
                except Exception:
                    ts = None
                bucket = ts.strftime("%Y-%m-%d %H") if ts else "unknown"
                key = (model_name, bucket)
                g = groups[key]
                g["count"] += 1
                if isinstance(row.get("score"), (int, float)):
                    g["scores"].append(float(row["score"]))
                if row.get("response_error"):
                    g["errors"] += 1
                if ts:
                    if g["first_ts"] is None or ts < g["first_ts"]:
                        g["first_ts"] = ts
                    if g["last_ts"] is None or ts > g["last_ts"]:
                        g["last_ts"] = ts
            for (model_name, bucket), g in groups.items():
                avg = round(sum(g["scores"]) / len(g["scores"]), 3) if g["scores"] else None
                out.append({
                    "challenge_uuid": ch_uuid,
                    "challenge_name": ch_name,
                    "challenge_task_type": ch_task,
                    "model_name": model_name,
                    "count": g["count"],
                    "errors": g["errors"],
                    "avg_score": avg,
                    "started_at": g["first_ts"].isoformat() if g["first_ts"] else None,
                    "finished_at": g["last_ts"].isoformat() if g["last_ts"] else None,
                })
        return out

    runs = await asyncio.to_thread(_parse_all)

    # Sort by finish time desc (latest first)
    runs.sort(key=lambda r: r["finished_at"] or "", reverse=True)
    result = {"total": len(runs), "runs": runs}

    # Cache for 60 seconds
    _test_runs_cache["data"] = result
    _test_runs_cache["expires_at"] = time.monotonic() + 60
    return result


@router.get("/benchmarks", response_model=List[BenchmarkListItem])
async def list_benchmarks(
    status: Optional[str] = Query(None, description="Filter by status"),
    limit: int = Query(20, ge=1, le=100),
    offset: int = Query(0, ge=0),
    db: AsyncSession = Depends(get_db),
):
    """List all benchmarks with optional filtering."""
    benchmarks = await BenchmarkCRUD.list_all(db, status=status, limit=limit, offset=offset)
    return benchmarks


@router.post("/benchmarks", response_model=BenchmarkResponse)
async def create_benchmark(
    background_tasks: BackgroundTasks,
    name: str = Form(...),
    description: Optional[str] = Form(None),
    config_file: Optional[UploadFile] = File(None),
    config_text: Optional[str] = Form(None),
    auto_run: bool = Form(True),
    db: AsyncSession = Depends(get_db),
):
    """Create a new benchmark from YAML config and optionally start execution.

    Args:
        background_tasks: FastAPI background tasks
        name: Benchmark name
        description: Optional description
        config_file: YAML config file upload
        config_text: YAML config as text
        auto_run: If True, automatically start benchmark execution (default: True)
        db: Database session
    """
    # Get config YAML from either file or text
    if config_file:
        config_yaml = (await config_file.read()).decode("utf-8")
    elif config_text:
        config_yaml = config_text
    else:
        raise HTTPException(status_code=400, detail="Must provide config_file or config_text")

    # Validate config
    try:
        from ...config.loader import load_config_from_string
        cfg = load_config_from_string(config_yaml)
        config_fingerprint = cfg.fingerprint()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid config: {str(e)}")

    # Extract metadata
    metadata = cfg.metadata if hasattr(cfg, "metadata") else {}
    metadata_json = json.dumps(metadata) if metadata else None

    # Create benchmark
    benchmark = await BenchmarkCRUD.create(
        session=db,
        name=name,
        description=description,
        config_yaml=config_yaml,
        config_fingerprint=config_fingerprint,
        metadata_json=metadata_json,
    )
    await db.commit()
    await db.refresh(benchmark)

    # Start benchmark execution in background if auto_run is True
    if auto_run:
        # Use asyncio.create_task instead of BackgroundTasks for better async support
        from ..task_manager import task_manager
        task = asyncio.create_task(run_benchmark_task(benchmark.uuid, config_yaml))
        task_manager.register_task(benchmark.uuid, task)

    # Return with empty scenarios list (benchmark is newly created)
    return BenchmarkResponse(
        uuid=benchmark.uuid,
        name=benchmark.name,
        description=benchmark.description,
        config_fingerprint=benchmark.config_fingerprint,
        status=benchmark.status,
        created_at=benchmark.created_at,
        started_at=benchmark.started_at,
        completed_at=benchmark.completed_at,
        runtime_sec=benchmark.runtime_sec,
        error_message=benchmark.error_message,
        scenarios=[],
    )


@router.get("/benchmarks/{uuid}", response_model=BenchmarkResponse)
async def get_benchmark(
    uuid: str,
    db: AsyncSession = Depends(get_db),
):
    """Get benchmark details with scenarios."""
    benchmark = await BenchmarkCRUD.get_with_scenarios(db, uuid)
    if not benchmark:
        raise HTTPException(status_code=404, detail="Benchmark not found")

    # Manually construct response to avoid async relationship access issues
    from ..schemas import ScenarioResponse
    scenarios = [
        ScenarioResponse(
            scenario_name=s.scenario_name,
            server_name=s.server_name,
            request_count=s.request_count,
            error_count=s.error_count,
            p50_ms=s.p50_ms,
            p95_ms=s.p95_ms,
            p99_ms=s.p99_ms,
            avg_ms=s.avg_ms,
            tokens_per_sec_output=s.tokens_per_sec_output,
            tokens_per_sec_total=s.tokens_per_sec_total,
            error_rate=s.error_rate,
        )
        for s in benchmark.scenarios
    ]

    return BenchmarkResponse(
        uuid=benchmark.uuid,
        name=benchmark.name,
        description=benchmark.description,
        config_fingerprint=benchmark.config_fingerprint,
        status=benchmark.status,
        created_at=benchmark.created_at,
        started_at=benchmark.started_at,
        completed_at=benchmark.completed_at,
        runtime_sec=benchmark.runtime_sec,
        error_message=benchmark.error_message,
        scenarios=scenarios,
    )


@router.get("/benchmarks/{uuid}/status", response_model=BenchmarkStatus)
async def get_benchmark_status(
    uuid: str,
    db: AsyncSession = Depends(get_db),
):
    """Get current benchmark status (for polling)."""
    benchmark = await BenchmarkCRUD.get_by_uuid(db, uuid)
    if not benchmark:
        raise HTTPException(status_code=404, detail="Benchmark not found")

    # Calculate progress if running
    progress = None
    current_request_count = None

    if benchmark.status == "running" and benchmark.scenarios:
        # Sum up request counts from scenarios
        current_request_count = sum(s.request_count or 0 for s in benchmark.scenarios)
        # Progress is approximate (we don't know total expected requests easily)
        progress = min(int((current_request_count / 100) * 100), 100) if current_request_count else 0

    return BenchmarkStatus(
        uuid=benchmark.uuid,
        status=benchmark.status,
        progress=progress,
        current_request_count=current_request_count,
    )


@router.post("/benchmarks/{uuid}/run")
async def run_benchmark_endpoint(
    uuid: str,
    db: AsyncSession = Depends(get_db),
):
    """Manually start a benchmark that was created with auto_run=False."""
    benchmark = await BenchmarkCRUD.get_by_uuid(db, uuid)
    if not benchmark:
        raise HTTPException(status_code=404, detail="Benchmark not found")

    if benchmark.status != "pending":
        raise HTTPException(
            status_code=400,
            detail=f"Can only run pending benchmarks (current status: {benchmark.status})"
        )

    # Start benchmark execution in background
    from ..task_manager import task_manager
    task = asyncio.create_task(run_benchmark_task(benchmark.uuid, benchmark.config_yaml))
    task_manager.register_task(benchmark.uuid, task)

    return JSONResponse({"message": "Benchmark started", "uuid": uuid})


@router.post("/benchmarks/{uuid}/cancel")
async def cancel_benchmark(
    uuid: str,
    db: AsyncSession = Depends(get_db),
):
    """Cancel a running benchmark."""
    from ..task_manager import task_manager
    from ..log_handler import add_log

    benchmark = await BenchmarkCRUD.get_by_uuid(db, uuid)
    if not benchmark:
        raise HTTPException(status_code=404, detail="Benchmark not found")

    if benchmark.status not in ["pending", "running"]:
        raise HTTPException(status_code=400, detail="Can only cancel pending or running benchmarks")

    # Request task cancellation
    cancelled = task_manager.cancel_task(uuid)

    # Log the cancellation request
    await add_log(db, benchmark.id, "WARNING", "Cancellation requested by user", "system")

    # Update status to cancelled
    await BenchmarkCRUD.update_status(db, uuid, status="cancelled")
    await db.commit()

    return {
        "message": "Benchmark cancellation requested" if cancelled else "Benchmark marked as cancelled",
        "uuid": uuid,
        "task_found": cancelled
    }


@router.delete("/benchmarks/{uuid}")
async def delete_benchmark(
    uuid: str,
    db: AsyncSession = Depends(get_db),
):
    """Delete a benchmark and all related data."""
    benchmark = await BenchmarkCRUD.get_by_uuid(db, uuid)
    if not benchmark:
        raise HTTPException(status_code=404, detail="Benchmark not found")

    await BenchmarkCRUD.delete(db, uuid)

    return {"message": "Benchmark deleted", "uuid": uuid}


@router.post("/benchmarks/{uuid}/rerun")
async def rerun_benchmark(
    uuid: str,
    auto_run: bool = Form(True),
    db: AsyncSession = Depends(get_db),
):
    """Create a new benchmark run with the same configuration."""
    from ..tasks import run_benchmark_task

    # Get the original benchmark
    original = await BenchmarkCRUD.get_by_uuid(db, uuid)
    if not original:
        raise HTTPException(status_code=404, detail="Benchmark not found")

    # Get the next run number
    next_run_number = await BenchmarkCRUD.get_next_run_number(db, uuid)

    # Determine the root parent
    root_parent_uuid = original.parent_uuid if original.parent_uuid else uuid

    # Create new benchmark with same config
    new_benchmark = await BenchmarkCRUD.create(
        session=db,
        name=f"{original.name} (Run #{next_run_number})",
        description=original.description,
        config_yaml=original.config_yaml,
        config_fingerprint=original.config_fingerprint,
        parent_uuid=root_parent_uuid,
        run_number=next_run_number,
    )
    await db.commit()

    # Start execution if auto_run is True
    if auto_run:
        from ..task_manager import task_manager
        task = asyncio.create_task(run_benchmark_task(new_benchmark.uuid, new_benchmark.config_yaml))
        task_manager.register_task(new_benchmark.uuid, task)

    return BenchmarkResponse(
        uuid=new_benchmark.uuid,
        name=new_benchmark.name,
        description=new_benchmark.description,
        status=new_benchmark.status,
        config_fingerprint=new_benchmark.config_fingerprint,
        created_at=new_benchmark.created_at,
        started_at=new_benchmark.started_at,
        completed_at=new_benchmark.completed_at,
        runtime_sec=new_benchmark.runtime_sec,
        error_message=new_benchmark.error_message,
        scenarios=[],
    )


@router.get("/benchmarks/{uuid}/history")
async def get_benchmark_history(
    uuid: str,
    db: AsyncSession = Depends(get_db),
):
    """Get all benchmarks in the same history chain."""
    history = await BenchmarkCRUD.get_history(db, uuid)

    from ..schemas import BenchmarkListItem
    return [
        BenchmarkListItem(
            uuid=b.uuid,
            name=b.name,
            description=b.description,
            status=b.status,
            created_at=b.created_at,
            runtime_sec=b.runtime_sec,
            error_message=b.error_message,
        )
        for b in history
    ]


@router.get("/benchmarks/{uuid}/config")
async def get_benchmark_config(
    uuid: str,
    db: AsyncSession = Depends(get_db),
):
    """Get the YAML configuration of a benchmark for editing."""
    benchmark = await BenchmarkCRUD.get_by_uuid(db, uuid)
    if not benchmark:
        raise HTTPException(status_code=404, detail="Benchmark not found")

    return {
        "uuid": benchmark.uuid,
        "name": benchmark.name,
        "description": benchmark.description,
        "config_yaml": benchmark.config_yaml,
        "config_fingerprint": benchmark.config_fingerprint,
    }


@router.get("/benchmarks/{uuid}/logs")
async def get_benchmark_logs(
    uuid: str,
    since_id: int = 0,
    limit: int = 100,
    db: AsyncSession = Depends(get_db),
):
    """Get execution logs for a benchmark.

    Args:
        uuid: Benchmark UUID
        since_id: Only return logs with ID > since_id (for polling)
        limit: Maximum number of logs to return
    """
    from sqlalchemy import select
    from ..models import BenchmarkLog

    # Get benchmark
    benchmark = await BenchmarkCRUD.get_by_uuid(db, uuid)
    if not benchmark:
        raise HTTPException(status_code=404, detail="Benchmark not found")

    # Query logs
    query = (
        select(BenchmarkLog)
        .where(BenchmarkLog.benchmark_id == benchmark.id)
        .where(BenchmarkLog.id > since_id)
        .order_by(BenchmarkLog.id.asc())
        .limit(limit)
    )

    result = await db.execute(query)
    logs = result.scalars().all()

    return [
        {
            "id": log.id,
            "timestamp": log.timestamp.isoformat(),
            "level": log.level,
            "message": log.message,
            "source": log.source,
        }
        for log in logs
    ]


@router.get("/benchmarks/{uuid}/export")
async def export_benchmark(
    uuid: str,
    format: str = Query("json", regex="^(json|csv|markdown)$"),
    db: AsyncSession = Depends(get_db),
):
    """Export benchmark results."""
    benchmark = await BenchmarkCRUD.get_with_scenarios(db, uuid)
    if not benchmark:
        raise HTTPException(status_code=404, detail="Benchmark not found")

    if format == "json":
        # Return full benchmark data as JSON
        data = {
            "benchmark": benchmark.to_dict(),
            "scenarios": [s.to_dict() for s in benchmark.scenarios],
        }
        return JSONResponse(content=data)

    elif format == "csv":
        # Generate CSV of scenario metrics
        import csv
        import io

        output = io.StringIO()
        writer = csv.DictWriter(
            output,
            fieldnames=["scenario_name", "server_name", "request_count", "p50_ms", "p95_ms", "tokens_per_sec_output", "error_rate"],
        )
        writer.writeheader()

        for scenario in benchmark.scenarios:
            writer.writerow({
                "scenario_name": scenario.scenario_name,
                "server_name": scenario.server_name,
                "request_count": scenario.request_count or 0,
                "p50_ms": scenario.p50_ms or 0,
                "p95_ms": scenario.p95_ms or 0,
                "tokens_per_sec_output": scenario.tokens_per_sec_output or 0,
                "error_rate": scenario.error_rate or 0,
            })

        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=benchmark_{uuid}.csv"},
        )

    elif format == "markdown":
        # Generate markdown report (simple version)
        lines = [
            f"# Benchmark: {benchmark.name}",
            "",
            f"**Status**: {benchmark.status}",
            f"**Created**: {benchmark.created_at}",
            f"**Runtime**: {benchmark.runtime_sec:.2f}s" if benchmark.runtime_sec else "**Runtime**: N/A",
            "",
            "## Scenarios",
            "",
        ]

        for scenario in benchmark.scenarios:
            lines.extend([
                f"### {scenario.scenario_name} - {scenario.server_name}",
                "",
                f"- Requests: {scenario.request_count or 0}",
                f"- p50: {scenario.p50_ms:.1f}ms" if scenario.p50_ms else "- p50: N/A",
                f"- p95: {scenario.p95_ms:.1f}ms" if scenario.p95_ms else "- p95: N/A",
                f"- Throughput: {scenario.tokens_per_sec_output:.2f} tokens/s" if scenario.tokens_per_sec_output else "- Throughput: N/A",
                f"- Error Rate: {scenario.error_rate:.3f}" if scenario.error_rate is not None else "- Error Rate: N/A",
                "",
            ])

        content = "\n".join(lines)
        return StreamingResponse(
            iter([content]),
            media_type="text/markdown",
            headers={"Content-Disposition": f"attachment; filename=benchmark_{uuid}.md"},
        )


@router.post("/validate-config")
async def validate_config(
    config_text: str = Form(...),
):
    """Validate a YAML config without creating a benchmark."""
    try:
        from ...config.loader import load_config_from_string
        cfg = load_config_from_string(config_text)
        return {
            "valid": True,
            "fingerprint": cfg.fingerprint(),
            "servers": [s.name for s in cfg.servers],
            "scenarios": [s.name for s in cfg.scenarios],
        }
    except Exception as e:
        return {
            "valid": False,
            "error": str(e),
        }


# Server management endpoints
@router.get("/servers", response_model=List[ServerResponse])
async def list_servers(db: AsyncSession = Depends(get_db)):
    """List all saved server configurations."""
    servers = await ServerCRUD.list_all(db)
    return servers


@router.post("/servers", response_model=ServerResponse)
async def create_server(
    server: ServerCreate,
    db: AsyncSession = Depends(get_db),
):
    """Create a new server configuration."""
    # Check if server with same name exists
    existing = await ServerCRUD.get_by_name(db, server.name)
    if existing:
        raise HTTPException(status_code=400, detail="Server with this name already exists")

    new_server = await ServerCRUD.create(
        session=db,
        name=server.name,
        type=server.type,
        base_url=server.base_url,
        model=server.model,
        config_json=server.config_json,
    )
    return new_server


@router.delete("/servers/{name}")
async def delete_server(
    name: str,
    db: AsyncSession = Depends(get_db),
):
    """Delete a server configuration."""
    server = await ServerCRUD.get_by_name(db, name)
    if not server:
        raise HTTPException(status_code=404, detail="Server not found")

    await ServerCRUD.delete(db, name)
    return {"message": "Server deleted", "name": name}


# Challenge endpoints
@router.get("/challenges", response_model=List[ChallengeListItem])
async def list_challenges(
    task_type: Optional[str] = Query(None),
    limit: int = Query(50, ge=1, le=200),
    offset: int = Query(0, ge=0),
    db: AsyncSession = Depends(get_db),
):
    """List all challenges."""
    return await ChallengeCRUD.list_all(db, task_type=task_type, limit=limit, offset=offset)


@router.post("/challenges", response_model=ChallengeResponse)
async def create_challenge(
    name: str = Form(...),
    description: Optional[str] = Form(None),
    task_type: Optional[str] = Form(None),
    data_file: Optional[UploadFile] = File(None),
    db: AsyncSession = Depends(get_db),
):
    """Create a new challenge, optionally with JSONL data."""
    data_jsonl = None
    if data_file:
        content = await data_file.read()
        data_jsonl = content.decode("utf-8")
        # Validate JSONL
        for i, line in enumerate(data_jsonl.splitlines()):
            line = line.strip()
            if not line:
                continue
            try:
                json.loads(line)
            except json.JSONDecodeError as e:
                raise HTTPException(status_code=400, detail=f"Invalid JSON on line {i+1}: {e}")

    challenge = await ChallengeCRUD.create(
        session=db, name=name, description=description,
        task_type=task_type, data_jsonl=data_jsonl,
    )
    await db.commit()
    await db.refresh(challenge)
    return challenge


@router.get("/challenges/{uuid}", response_model=ChallengeResponse)
async def get_challenge(uuid: str, db: AsyncSession = Depends(get_db)):
    """Get challenge details."""
    challenge = await ChallengeCRUD.get_by_uuid(db, uuid)
    if not challenge:
        raise HTTPException(status_code=404, detail="Challenge not found")
    return challenge


@router.get("/challenges/{uuid}/preview")
async def preview_challenge(
    uuid: str,
    n: int = Query(5, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
):
    """Preview first N rows of challenge data."""
    challenge = await ChallengeCRUD.get_by_uuid(db, uuid)
    if not challenge:
        raise HTTPException(status_code=404, detail="Challenge not found")

    rows = []
    if challenge.data_jsonl:
        for line in challenge.data_jsonl.splitlines():
            line = line.strip()
            if not line:
                continue
            rows.append(json.loads(line))
            if len(rows) >= n:
                break

    return {"uuid": uuid, "preview_count": len(rows), "total_rows": challenge.row_count, "rows": rows}


@router.get("/challenges/{uuid}/aspect-stats")
async def challenge_aspect_stats(
    uuid: str,
    aspects: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
):
    """Return source-material counts per aspect/subject before generation."""
    challenge = await ChallengeCRUD.get_by_uuid(db, uuid)
    if not challenge:
        raise HTTPException(status_code=404, detail="Challenge not found")

    if not challenge.data_jsonl:
        return {
            "total_rows": 0,
            "matched_rows": 0,
            "selected_aspects": list(_DEFAULT_ASPECT_LABELS),
            "aspect_counts": {label: 0 for label in _DEFAULT_ASPECT_LABELS},
        }

    selected = _parse_selected_aspects(aspects)
    filtered_jsonl, total_rows, matched_rows = _filter_jsonl_by_aspects(
        challenge.data_jsonl, selected
    )
    counts = _compute_aspect_stats(filtered_jsonl)

    return {
        "total_rows": total_rows,
        "matched_rows": matched_rows,
        "selected_aspects": selected,
        "aspect_counts": counts,
    }


@router.put("/challenges/{uuid}", response_model=ChallengeResponse)
async def update_challenge(
    uuid: str,
    name: str = Form(...),
    description: Optional[str] = Form(None),
    task_type: Optional[str] = Form(None),
    data_file: Optional[UploadFile] = File(None),
    db: AsyncSession = Depends(get_db),
):
    """Update a challenge."""
    data_jsonl = None
    if data_file and data_file.filename:
        content = await data_file.read()
        data_jsonl = content.decode("utf-8")
        for i, line in enumerate(data_jsonl.splitlines()):
            line = line.strip()
            if not line:
                continue
            try:
                json.loads(line)
            except json.JSONDecodeError as e:
                raise HTTPException(status_code=400, detail=f"Invalid JSON on line {i+1}: {e}")

    challenge = await ChallengeCRUD.update(
        session=db, uuid=uuid, name=name, description=description,
        task_type=task_type, data_jsonl=data_jsonl,
    )
    if not challenge:
        raise HTTPException(status_code=404, detail="Challenge not found")
    await db.commit()
    await db.refresh(challenge)
    return challenge


@router.post("/challenges/import/threads")
async def import_threads_challenge(
    name: str = Form(...),
    description: Optional[str] = Form(None),
    directory: str = Form(...),
    keyword: str = Form(""),
    include_replies: bool = Form(False),
    combine_replies: bool = Form(False),
    min_like_count: int = Form(0),
    min_replies_count: int = Form(0),
    min_repost_count: int = Form(0),
    date_start: Optional[str] = Form(None),
    date_end: Optional[str] = Form(None),
    text_contains: Optional[str] = Form(None),
    min_text_length: int = Form(0),
    exclude_emoji_only: bool = Form(False),
    limit: Optional[int] = Form(None),
    db: AsyncSession = Depends(get_db),
):
    """Import Threads scraper JSON files as a challenge."""
    import asyncio
    from pathlib import Path
    from ...qual.threads_source import load_threads_materials

    dir_path = Path(directory.strip().strip('"').strip("'"))
    if not dir_path.exists():
        raise HTTPException(status_code=400, detail=f"Directory not found: {directory}")

    loop = asyncio.get_event_loop()
    try:
        materials = await loop.run_in_executor(
            None,
            lambda: load_threads_materials(
                directory=dir_path,
                keyword=keyword,
                include_replies=include_replies,
                combine_replies=combine_replies,
                min_like_count=min_like_count,
                min_replies_count=min_replies_count,
                min_repost_count=min_repost_count,
                date_start=date_start or None,
                date_end=date_end or None,
                text_contains=text_contains or None,
                min_text_length=min_text_length,
                exclude_emoji_only=exclude_emoji_only,
                limit=limit,
            ),
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to load Threads data: {e}")

    if not materials:
        raise HTTPException(status_code=400, detail="No data found with the given filters.")

    # Convert RawMaterial to JSONL
    lines = []
    for m in materials:
        lines.append(json.dumps({
            "text": m.content,
            "title": m.title,
            "source_category": m.source_category,
            "keyword": m.keyword,
            "month": m.month_range.get("start", ""),
        }, ensure_ascii=False))
    data_jsonl = "\n".join(lines)

    challenge = await ChallengeCRUD.create(
        session=db,
        name=name,
        description=description,
        task_type="threads",
        data_jsonl=data_jsonl,
    )
    await db.commit()
    await db.refresh(challenge)
    return challenge


@router.post("/challenges/import/taiwan-md")
async def import_taiwan_md_challenge(
    name: str = Form(...),
    description: Optional[str] = Form(None),
    categories: Optional[str] = Form(None),
    lang: str = Form("zh-TW"),
    limit: Optional[int] = Form(None),
    db: AsyncSession = Depends(get_db),
):
    """Import Taiwan.md articles as a challenge."""
    import asyncio
    from ...qual.taiwan_md_source import load_taiwan_md_materials

    cat_list = [c.strip() for c in categories.split(",")] if categories else None

    loop = asyncio.get_event_loop()
    try:
        materials = await loop.run_in_executor(
            None,
            lambda: load_taiwan_md_materials(
                categories=cat_list,
                lang=lang,
                limit=limit,
            ),
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to load Taiwan.md data: {e}")

    if not materials:
        raise HTTPException(status_code=400, detail="No articles found with the given filters.")

    lines = []
    for m in materials:
        lines.append(json.dumps({
            "text": m.content,
            "title": m.title,
            "source_category": m.source_category,
            "keyword": m.keyword,
        }, ensure_ascii=False))
    data_jsonl = "\n".join(lines)

    challenge = await ChallengeCRUD.create(
        session=db,
        name=name,
        description=description,
        task_type="taiwan_md",
        data_jsonl=data_jsonl,
    )
    await db.commit()
    await db.refresh(challenge)
    return challenge


@router.post("/challenges/import/taiwan-knowledge")
async def import_taiwan_knowledge_challenge(
    name: str = Form(...),
    description: Optional[str] = Form(None),
    titles: Optional[str] = Form(None),
    include_world_knowledge: str = Form("0"),
    limit: Optional[int] = Form(None),
    db: AsyncSession = Depends(get_db),
):
    """Import Taiwan Wikipedia knowledge articles as a challenge.

    When include_world_knowledge is "1", also fetches foreign-country articles
    and appends them so the Designer can generate cross-country confusion tests.
    """
    import asyncio
    from ...qual.taiwan_knowledge_source import load_taiwan_knowledge_materials
    from ...qual.world_knowledge_source import load_world_knowledge_materials

    title_list = [t.strip() for t in titles.split(",")] if titles else None

    loop = asyncio.get_event_loop()
    try:
        materials = await loop.run_in_executor(
            None,
            lambda: load_taiwan_knowledge_materials(
                titles=title_list,
                max_chars=3000,
            ),
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to load Taiwan knowledge data: {e}")

    if include_world_knowledge == "1":
        try:
            world_materials = await loop.run_in_executor(
                None,
                lambda: load_world_knowledge_materials(max_chars=3000),
            )
            materials = materials + world_materials
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to load world knowledge data: {e}")

    if not materials:
        raise HTTPException(status_code=400, detail="No articles found.")

    lines = []
    for m in materials:
        lines.append(json.dumps({
            "text": m.content,
            "title": m.title,
            "source_category": m.source_category,
            "keyword": m.keyword,
        }, ensure_ascii=False))
    data_jsonl = "\n".join(lines)

    challenge = await ChallengeCRUD.create(
        session=db,
        name=name,
        description=description,
        task_type="true_false",
        data_jsonl=data_jsonl,
    )
    await db.commit()
    await db.refresh(challenge)
    return challenge


@router.post("/challenges/import/world-knowledge")
async def import_world_knowledge_challenge(
    name: str = Form(...),
    description: Optional[str] = Form(None),
    titles: Optional[str] = Form(None),
    limit: Optional[int] = Form(None),
    db: AsyncSession = Depends(get_db),
):
    """Import foreign-country Wikipedia articles as a challenge for cross-country confusion tests."""
    import asyncio
    from ...qual.world_knowledge_source import load_world_knowledge_materials

    title_list = [t.strip() for t in titles.split(",")] if titles else None

    loop = asyncio.get_event_loop()
    try:
        materials = await loop.run_in_executor(
            None,
            lambda: load_world_knowledge_materials(
                titles=title_list,
                max_chars=3000,
            ),
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to load world knowledge data: {e}")

    if not materials:
        raise HTTPException(status_code=400, detail="No articles found.")

    if limit:
        materials = materials[:limit]

    lines = []
    for m in materials:
        lines.append(json.dumps({
            "text": m.content,
            "title": m.title,
            "source_category": m.source_category,
            "keyword": m.keyword,
        }, ensure_ascii=False))
    data_jsonl = "\n".join(lines)

    challenge = await ChallengeCRUD.create(
        session=db,
        name=name,
        description=description,
        task_type="true_false",
        data_jsonl=data_jsonl,
    )
    await db.commit()
    await db.refresh(challenge)
    return challenge


@router.post("/challenges/import/school-exam")
async def import_school_exam_challenge(
    name: str = Form(...),
    description: Optional[str] = Form(None),
    source: str = Form("builtin"),          # "builtin" or "exambank"
    level: str = Form("both"),
    subjects: Optional[str] = Form(None),   # comma-separated
    grades: Optional[str] = Form(None),     # comma-separated, exambank only
    manifest: Optional[str] = Form(None),   # exambank only
    zip_archives: Optional[str] = Form(None),  # comma-separated paths, exambank only
    parse_questions: str = Form("0"),           # "1" = direct MCQ parse, exambank only
    limit: Optional[int] = Form(None),
    db: AsyncSession = Depends(get_db),
):
    """Import school exam / curriculum materials as a challenge."""
    import asyncio

    subject_list = [s.strip() for s in subjects.split(",")] if subjects else None
    grade_list   = [g.strip() for g in grades.split(",")]   if grades   else None
    zip_list     = _parse_path_list(zip_archives) if zip_archives else None

    loop = asyncio.get_event_loop()
    try:
        if source == "exambank":
            from ...qual.exam_bank_source import load_exam_bank_materials
            _parse_questions = parse_questions == "1"
            materials = await loop.run_in_executor(
                None,
                lambda: load_exam_bank_materials(
                    manifest=manifest or None,
                    zip_archives=zip_list,
                    level=level,
                    subjects=subject_list,
                    grades=grade_list,
                    limit=limit,
                    parse_questions=_parse_questions,
                ),
            )
        else:
            from ...qual.school_qa_source import load_school_qa_materials
            materials = await loop.run_in_executor(
                None,
                lambda: load_school_qa_materials(
                    level=level,
                    subjects=subject_list,
                    limit=limit,
                ),
            )
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"匯入失敗：{e}")

    if not materials:
        raise HTTPException(status_code=400, detail="沒有找到符合條件的材料，請放寬篩選條件。")

    lines = []
    for m in materials:
        lines.append(json.dumps({
            "text": m.content,
            "title": m.title,
            "source_category": m.source_category,
            "keyword": m.keyword,
        }, ensure_ascii=False))
    data_jsonl = "\n".join(lines)

    challenge = await ChallengeCRUD.create(
        session=db,
        name=name,
        description=description,
        task_type="school_qa",
        data_jsonl=data_jsonl,
    )
    await db.commit()
    await db.refresh(challenge)
    return challenge


@router.post("/challenges/import/ptt-movie")
async def import_ptt_movie_challenge(
    name: str = Form(...),
    description: Optional[str] = Form(None),
    board: str = Form("movie"),
    pages: int = Form(1),
    keyword: str = Form("movie"),
    title_contains: Optional[str] = Form(None),
    text_contains: Optional[str] = Form(None),
    min_push_count: int = Form(0),
    date_start: Optional[str] = Form(None),
    date_end: Optional[str] = Form(None),
    combine_pushes: bool = Form(True),
    include_pushes: bool = Form(False),
    max_pushes_per_article: int = Form(10),
    min_text_length: int = Form(0),
    exclude_re_posts: bool = Form(True),
    exclude_fw_posts: bool = Form(True),
    exclude_announcements: bool = Form(True),
    limit: Optional[int] = Form(None),
    db: AsyncSession = Depends(get_db),
):
    """Import PTT board articles as a challenge."""
    import asyncio
    from ...qual.ptt_source import PTTLoadReport, load_ptt_board_materials

    exclude_prefixes: list[str] = []
    if exclude_re_posts:
        exclude_prefixes.append("Re:")
    if exclude_fw_posts:
        exclude_prefixes.append("Fw:")
    if exclude_announcements:
        exclude_prefixes.extend(["[公告]", "公告"])

    loop = asyncio.get_event_loop()
    report = PTTLoadReport(pages_requested=pages)
    try:
        materials, report = await loop.run_in_executor(
            None,
            lambda: load_ptt_board_materials(
                board=board,
                pages=pages,
                keyword=keyword,
                title_contains=title_contains or None,
                text_contains=text_contains or None,
                min_push_count=min_push_count,
                date_start=date_start or None,
                date_end=date_end or None,
                combine_pushes=combine_pushes,
                include_pushes=include_pushes,
                max_pushes_per_article=max_pushes_per_article,
                min_text_length=min_text_length,
                exclude_title_prefixes=exclude_prefixes,
                limit=limit,
                return_report=True,
            ),
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to load PTT data: {e}")

    if not materials:
        raise HTTPException(status_code=400, detail="No PTT articles found with the given filters.")

    lines = []
    for m in materials:
        lines.append(json.dumps({
            "text": m.content,
            "title": m.title,
            "source_category": m.source_category,
            "keyword": m.keyword,
            "month": m.month_range.get("start", ""),
        }, ensure_ascii=False))
    data_jsonl = "\n".join(lines)

    challenge = await ChallengeCRUD.create(
        session=db,
        name=name,
        description=description,
        task_type=f"ptt_{board.strip().lower()}",
        data_jsonl=data_jsonl,
    )
    await db.commit()
    await db.refresh(challenge)
    return {
        "uuid": challenge.uuid,
        "name": challenge.name,
        "description": challenge.description,
        "task_type": challenge.task_type,
        "row_count": challenge.row_count,
        "created_at": challenge.created_at,
        "updated_at": challenge.updated_at,
        "import_summary": report.to_dict(),
    }


# Track in-progress generate tasks: uuid → asyncio.Task
_running_generates: dict[str, asyncio.Task] = {}
# Track pipeline errors: uuid → error message
_generate_errors: dict[str, str] = {}


async def _run_generate_background(uuid: str, config, parsed_model_names, selected_aspects,
                                    normalized_target, sample_size, normalized_scope,
                                    source_row_count, matched_source_row_count, sampled_row_count,
                                    effective_model_base_url, effective_judge_model_name) -> None:
    """Run the qual pipeline in background and save results when done."""
    import os
    from ...qual.pipeline import run_qual_pipeline

    _generate_errors.pop(uuid, None)
    try:
        result = await run_qual_pipeline(config)
    except asyncio.CancelledError:
        logger.info("generate: task cancelled for challenge %s", uuid)
        return
    except Exception as exc:
        logger.error("generate: pipeline error for challenge %s: %s", uuid, exc)
        _generate_errors[uuid] = str(exc)
        return
    finally:
        _running_generates.pop(uuid, None)

    # Build item rows and persist results
    item_rows = []
    results_lines = []
    filtered_count = 0
    score_by_item_model = {(s.benchmark_item_id, s.model_name): s for s in result.scores}
    response_by_item_model = {(r.benchmark_item_id, r.model_name): r for r in result.responses}

    # Within-run dedup for true_false: filter items first so we know the final dedup count.
    seen_statements_this_run: set[str] = set()
    dedup_dropped_count = 0
    filtered_items_list = []
    for item in result.dataset.items:
        if item.task_type.value == "true_false":
            try:
                _ref_obj = json.loads(item.reference_answer) if item.reference_answer else {}
                _stmt = _ref_obj.get("statement", "") if isinstance(_ref_obj, dict) else ""
            except Exception:
                _stmt = ""
            if _stmt:
                if _stmt in seen_statements_this_run:
                    dedup_dropped_count += 1
                    continue
                seen_statements_this_run.add(_stmt)
        filtered_items_list.append(item)

    for item in filtered_items_list:
        topic = ""
        thread_post = ""
        thread_replies = ""
        for tested_model in parsed_model_names:
            response = response_by_item_model.get((item.id, tested_model))
            judge = score_by_item_model.get((item.id, tested_model))
            rubric = item.scoring_rubric or ""
            stance = ""
            row = {
                "task_type": item.task_type.value,
                "model_name": tested_model,
                "judge_model": effective_judge_model_name,
                "leaderboard_target": normalized_target,
                "leaderboard_sample_size": sample_size,
                "leaderboard_sample_scope": normalized_scope,
                "model_api_base_url": effective_model_base_url,
                "selected_aspects": selected_aspects,
                "topic": topic,
                "stance": stance,
                "title": item.source_material.title or "",
                "keyword": item.source_material.keyword or "",
                "content": item.source_material.content,
                "full_content": item.source_material.content,
                "aspect": _infer_aspect_from_text(
                    item.source_material.title,
                    item.source_material.content,
                    item.source_material.keyword,
                    item.source_material.source_category,
                ),
                "prompt": item.prompt,
                "thread_post": thread_post,
                "thread_replies": thread_replies,
                "reference_answer": item.reference_answer or "",
                "scoring_rubric": rubric,
                "score": judge.score if judge else None,
                "reasoning": judge.reasoning if judge else "",
                "response_text": response.response_text if response else "",
                "response_error": response.error if response else "",
                "request_messages": response.request_messages if response else [],
                "requested_model": response.requested_model if response else tested_model,
                "requested_at": response.requested_at if response else None,
                "finished_at": response.finished_at if response else None,
                "retry_count": response.retry_count if response else 0,
                "latency_ms": round(response.latency_ms, 1) if response else None,
                "token_count": response.token_count if response else None,
                "_run_dedup_count": dedup_dropped_count,
            }
            item_rows.append(row)
            results_lines.append(json.dumps(row, ensure_ascii=False))

    results_jsonl = "\n".join(results_lines)
    async with AsyncSessionLocal() as db:
        await ChallengeCRUD.save_results(db, uuid, results_jsonl)
        await db.commit()
    logger.info("generate: saved %d result rows for challenge %s", len(item_rows), uuid)


@router.post("/challenges/{uuid}/generate")
async def generate_from_challenge(
    uuid: str,
    model_name: str = Form("gpt-oss-20b"),
    model_names: Optional[str] = Form(None),
    model_api_base_url: Optional[str] = Form(None),
    judge_model_name: Optional[str] = Form(None),
    task_types: str = Form("summarization,sentiment,classification,qa"),
    aspects: Optional[str] = Form(None),
    leaderboard_target: str = Form("model"),
    leaderboard_sample_size: int = Form(10),
    leaderboard_sample_scope: str = Form("overall"),
    items_per_task: int = Form(5),
    questions_per_article: int = Form(10),
    db: AsyncSession = Depends(get_db),
):
    """Run qual pipeline using challenge data as source material."""
    import asyncio
    import os
    from ...qual.config import (
        QualConfig, ChallengeSourceConfig, DataSourceConfig,
        ModelConfig, JudgeConfig,
    )
    from ...qual.pipeline import run_qual_pipeline

    challenge = await ChallengeCRUD.get_by_uuid(db, uuid)
    if not challenge:
        raise HTTPException(status_code=404, detail="Challenge not found")
    if not challenge.data_jsonl:
        raise HTTPException(status_code=400, detail="Challenge has no data")

    normalized_target = "participant" if leaderboard_target == "participant" else "model"
    normalized_scope = "per_aspect" if leaderboard_sample_scope == "per_aspect" else "overall"

    selected_aspects = _parse_selected_aspects(aspects)
    filtered_jsonl, source_row_count, matched_source_row_count = _filter_jsonl_by_aspects(
        challenge.data_jsonl, selected_aspects
    )
    if not filtered_jsonl.strip():
        raise HTTPException(
            status_code=400,
            detail=f"選定面向沒有符合的來源資料：{', '.join(selected_aspects)}",
        )

    # Source sampling cap: use the larger of leaderboard_sample_size and items_per_task,
    # bounded by the actual number of matching source articles.
    desired = max(int(leaderboard_sample_size or 10), int(items_per_task or 0))
    sample_size = max(1, min(desired, matched_source_row_count))
    sampled_jsonl, sampled_row_count = _sample_jsonl_rows(
        filtered_jsonl,
        selected_aspects,
        sample_size=sample_size,
        sample_scope=normalized_scope,
    )
    if not sampled_jsonl.strip():
        raise HTTPException(status_code=400, detail="抽樣後沒有可用題目。")

    litellm_url = os.getenv("LITELLM_URL", "https://llmgw.elandai.cloud")
    api_key = os.getenv("LITELLM_API_KEY", "")
    judge_api_key = os.getenv("JUDGE_LITELLM_API_KEY", api_key)
    effective_judge_model_name = (judge_model_name or os.getenv("JUDGE_MODEL_NAME") or model_name).strip()
    effective_model_base_url = (model_api_base_url or f"{litellm_url}/v1").strip()
    raw_model_names = model_names if model_names and model_names.strip() else model_name
    parsed_model_names = []
    seen_model_names: set[str] = set()
    for part in re.split(r"[\n,]+", raw_model_names):
        normalized = part.strip()
        if not normalized or normalized in seen_model_names:
            continue
        seen_model_names.add(normalized)
        parsed_model_names.append(normalized)
    if not parsed_model_names:
        raise HTTPException(status_code=400, detail="At least one model name is required")

    model_cfgs = [
        ModelConfig(
            name=model,
            base_url=effective_model_base_url,
            api_key=api_key,
            model=model,
        )
        for model in parsed_model_names
    ]
    judge_model_cfg = ModelConfig(
        name="judge",
        base_url=f"{litellm_url}/v1",
        api_key=judge_api_key,
        model=effective_judge_model_name,
    )
    config = QualConfig(
        data_source=DataSourceConfig(
            challenge=ChallengeSourceConfig(
                data_jsonl=sampled_jsonl,
                keyword=challenge.name,
            ),
        ),
        task_types=[t.strip() for t in task_types.split(",") if t.strip()],
        items_per_task=items_per_task,
        questions_per_article=max(1, min(int(questions_per_article or 10), 20)),
        models_under_test=model_cfgs,
        judge=JudgeConfig(model=judge_model_cfg, max_concurrent=2),
    )

    # Cancel any existing task for this challenge
    existing = _running_generates.pop(uuid, None)
    if existing and not existing.done():
        existing.cancel()

    task = asyncio.create_task(_run_generate_background(
        uuid=uuid, config=config, parsed_model_names=parsed_model_names,
        selected_aspects=selected_aspects, normalized_target=normalized_target,
        sample_size=sample_size, normalized_scope=normalized_scope,
        source_row_count=source_row_count, matched_source_row_count=matched_source_row_count,
        sampled_row_count=sampled_row_count, effective_model_base_url=effective_model_base_url,
        effective_judge_model_name=effective_judge_model_name,
    ))
    _running_generates[uuid] = task
    return {"status": "running", "message": f"生成已開始，共 {sampled_row_count} 筆材料"}


@router.get("/challenges/{uuid}/generate-status")
async def get_generate_status(uuid: str):
    """Return whether a background generate task is running for this challenge."""
    task = _running_generates.get(uuid)
    if task is None or task.done():
        error = _generate_errors.pop(uuid, None)
        if error:
            return {"status": "error", "error": error}
        return {"status": "idle"}
    return {"status": "running"}


@router.delete("/challenges/{uuid}/generate")
async def cancel_generate(uuid: str):
    """Cancel an in-progress background generate task."""
    task = _running_generates.pop(uuid, None)
    if task is None or task.done():
        return {"status": "idle", "message": "沒有進行中的生成任務"}
    task.cancel()
    return {"status": "cancelled", "message": "生成任務已取消"}


async def _run_test_new_model_background(
    uuid: str, model_name: str, judge_model_name: str,
    base_url: str, api_key: str, judge_api_key: str,
) -> None:
    """Run an existing question set against a new model and append to results_jsonl."""
    import os
    from openai import AsyncOpenAI, RateLimitError
    from ...qual.prompts.judge import JUDGE_SYSTEM_PROMPT, JUDGE_SCORE_PROMPT
    from datetime import datetime, timezone

    _generate_errors.pop(uuid, None)
    try:
        async with AsyncSessionLocal() as db:
            challenge = await ChallengeCRUD.get_by_uuid(db, uuid)
            if not challenge or not challenge.results_jsonl:
                _generate_errors[uuid] = "找不到既有題目"
                return

            # First pass: identify questions this model has SUCCESSFULLY answered.
            # A successful row has model_name match AND no response_error AND non-empty response_text.
            answered_keys: set[tuple] = set()
            kept_lines: list[str] = []
            for line in challenge.results_jsonl.splitlines():
                line = line.strip()
                if not line:
                    continue
                try:
                    row = json.loads(line)
                except Exception:
                    kept_lines.append(line)
                    continue
                row_model = row.get("model_name") or ""
                key = (row.get("prompt", ""), row.get("reference_answer", ""))
                if row_model == model_name:
                    is_success = (not row.get("response_error")) and bool(row.get("response_text"))
                    if is_success:
                        answered_keys.add(key)
                        kept_lines.append(line)
                    # else: failed row → drop it (will be retried)
                else:
                    kept_lines.append(line)

            # Now collect unique questions that THIS model hasn't successfully answered yet.
            seen: set[tuple] = set()
            unique_questions: list[dict] = []
            for line in kept_lines:
                try:
                    row = json.loads(line)
                except Exception:
                    continue
                key = (row.get("prompt", ""), row.get("reference_answer", ""))
                if not key[0] or key in seen or key in answered_keys:
                    continue
                seen.add(key)
                unique_questions.append(row)

            if not unique_questions:
                _generate_errors[uuid] = "沒有可測試的題目"
                return

            # Persist the cleaned results_jsonl (failed rows for this model dropped).
            existing_results = "\n".join(kept_lines)

        # Long timeout for reasoning models (gpt-5.x can take several minutes per question).
        client = AsyncOpenAI(base_url=base_url, api_key=api_key or "no-key", timeout=600.0)
        judge_client = AsyncOpenAI(base_url=base_url, api_key=judge_api_key or "no-key", timeout=120.0)
        sem = asyncio.Semaphore(4)

        async def run_one(ref_row: dict) -> dict | None:
            prompt = ref_row.get("prompt", "")
            ref = ref_row.get("reference_answer", "")
            rubric = ref_row.get("scoring_rubric", "")
            task_type = ref_row.get("task_type", "qa")

            requested_at = datetime.now(timezone.utc).isoformat()
            response_text = ""
            response_error = ""
            t0 = asyncio.get_event_loop().time()

            # Some LiteLLM gateways load-balance across deployments where only some
            # have a valid API key — retry on 401 hoping for a working deployment.
            # Also handle temperature unsupported errors for reasoning models.
            use_temperature = True
            max_auth_retries = 6

            async def _call():
                kwargs = {
                    "model": model_name,
                    "messages": [{"role": "user", "content": prompt}],
                }
                if use_temperature:
                    kwargs["temperature"] = 0.0
                return await client.chat.completions.create(**kwargs)

            async with sem:
                for attempt in range(max_auth_retries + 1):
                    try:
                        r = await _call()
                        response_text = (r.choices[0].message.content or "").strip()
                        response_error = ""
                        break
                    except Exception as e:
                        err_str = str(e)
                        response_error = err_str
                        # Temperature unsupported → drop it and retry immediately (no auth-retry counter)
                        if use_temperature and "temperature" in err_str and (
                            "does not support" in err_str or "Unsupported" in err_str
                        ):
                            use_temperature = False
                            continue
                        # Auth (401) → retry, LiteLLM will route to another deployment
                        is_auth = ("401" in err_str) or ("AuthenticationError" in err_str) or ("Incorrect API key" in err_str)
                        if is_auth and attempt < max_auth_retries:
                            await asyncio.sleep(0.5)
                            continue
                        # Other errors or out of retries → give up
                        break
            latency_ms = round((asyncio.get_event_loop().time() - t0) * 1000.0, 1)
            finished_at = datetime.now(timezone.utc).isoformat()

            # Judge the response
            score = None
            reasoning = ""
            if response_text and not response_error:
                judge_prompt = JUDGE_SCORE_PROMPT.format(
                    task_type=task_type, prompt=prompt,
                    reference_answer=ref, response=response_text, scoring_rubric=rubric,
                )
                async def _judge_call(use_temperature: bool):
                    kwargs = {
                        "model": judge_model_name,
                        "messages": [
                            {"role": "system", "content": JUDGE_SYSTEM_PROMPT},
                            {"role": "user", "content": judge_prompt},
                        ],
                    }
                    if use_temperature:
                        kwargs["temperature"] = 0.0
                    return await judge_client.chat.completions.create(**kwargs)

                try:
                    async with sem:
                        try:
                            jr = await _judge_call(use_temperature=True)
                        except Exception as je:
                            if "temperature" in str(je) and ("does not support" in str(je) or "Unsupported" in str(je)):
                                jr = await _judge_call(use_temperature=False)
                            else:
                                raise
                    jt = (jr.choices[0].message.content or "").strip()
                    if jt.startswith("```"):
                        lines = jt.splitlines()
                        lines = lines[1:] if lines[0].startswith("```") else lines
                        lines = lines[:-1] if lines and lines[-1].strip() == "```" else lines
                        jt = "\n".join(lines).strip()
                    jd = json.loads(jt)
                    score = jd.get("score")
                    reasoning = jd.get("reasoning", "")
                except Exception as e:
                    reasoning = f"評分失敗：{e}"

            new_row = dict(ref_row)
            new_row.update({
                "model_name": model_name,
                "judge_model": judge_model_name,
                "score": score,
                "reasoning": reasoning,
                "response_text": response_text,
                "response_error": response_error,
                "requested_model": model_name,
                "requested_at": requested_at,
                "finished_at": finished_at,
                "retry_count": 0,
                "latency_ms": latency_ms,
                "token_count": None,
            })
            return new_row

        tasks = [run_one(q) for q in unique_questions]
        new_rows = []
        for fut in asyncio.as_completed(tasks):
            try:
                row = await fut
                if row:
                    new_rows.append(row)
            except asyncio.CancelledError:
                raise
            except Exception:
                continue

        new_lines = [json.dumps(r, ensure_ascii=False) for r in new_rows]
        combined = (existing_results + "\n" + "\n".join(new_lines)).strip()

        async with AsyncSessionLocal() as db:
            await ChallengeCRUD.save_results(db, uuid, combined)
            await db.commit()
        logger.info("test-new-model: saved %d new rows for model %s on challenge %s",
                    len(new_rows), model_name, uuid)
    except asyncio.CancelledError:
        logger.info("test-new-model: task cancelled for challenge %s", uuid)
        return
    except Exception as exc:
        logger.error("test-new-model: error for challenge %s: %s", uuid, exc)
        _generate_errors[uuid] = str(exc)
    finally:
        _running_generates.pop(uuid, None)


@router.post("/challenges/{uuid}/test-new-model")
async def test_new_model(
    uuid: str,
    model_name: str = Form(...),
    judge_model_name: Optional[str] = Form(None),
    db: AsyncSession = Depends(get_db),
):
    """Test a new model against the existing question set without regenerating questions."""
    import os
    challenge = await ChallengeCRUD.get_by_uuid(db, uuid)
    if not challenge:
        raise HTTPException(status_code=404, detail="Challenge not found")
    if not challenge.results_jsonl:
        raise HTTPException(status_code=400, detail="尚未生成題目，請先執行 Generate。")

    existing = _running_generates.pop(uuid, None)
    if existing and not existing.done():
        existing.cancel()

    litellm_url = os.getenv("LITELLM_URL", "https://llmgw.elandai.cloud")
    base_url = f"{litellm_url}/v1"
    api_key = os.getenv("LITELLM_API_KEY", "")
    judge_api_key = os.getenv("JUDGE_LITELLM_API_KEY", api_key)
    effective_judge = (judge_model_name or os.getenv("JUDGE_MODEL_NAME") or "gpt-4o-mini").strip()

    task = asyncio.create_task(_run_test_new_model_background(
        uuid=uuid, model_name=model_name.strip(),
        judge_model_name=effective_judge,
        base_url=base_url, api_key=api_key, judge_api_key=judge_api_key,
    ))
    _running_generates[uuid] = task
    return {"status": "running", "message": f"開始用 {model_name} 測試現有題目"}


# Per-challenge results cache (60s TTL). Keyed by (uuid, results_jsonl hash).
# Hash of jsonl invalidates cache automatically when challenge data changes.
_results_cache: dict[str, dict] = {}


@router.get("/challenges/{uuid}/results")
async def get_challenge_results(uuid: str, db: AsyncSession = Depends(get_db)):
    """Get saved generated results for a challenge."""
    challenge = await ChallengeCRUD.get_by_uuid(db, uuid)
    if not challenge:
        raise HTTPException(status_code=404, detail="Challenge not found")
    default_labels = _DEFAULT_ASPECT_LABELS

    # Cache key includes a content fingerprint so re-runs invalidate automatically
    cache_key = uuid
    rj = challenge.results_jsonl or ""
    fingerprint = (len(rj), hash(rj[:1000]) if rj else 0)
    cached = _results_cache.get(cache_key)
    if cached and cached.get("fingerprint") == fingerprint and cached.get("expires_at", 0) > time.monotonic():
        return cached["data"]

    if not challenge.results_jsonl:
        return {
            "items": [],
            "total_items": 0,
            "source_items": 0,
            "filtered_items": 0,
            "filtered_reasons": {},
            "avg_score": None,
            "pass": None,
            "aspect_labels": list(default_labels),
            "selected_aspects": list(default_labels),
            "leaderboard_target": "model",
            "leaderboard_sample_size": 10,
            "leaderboard_sample_scope": "overall",
            "model_api_base_url": None,
            "tested_models": [],
            "judge_model_name": None,
            "per_model": {},
            "leaderboard": [],
        }
    raw_items = []
    for line in challenge.results_jsonl.splitlines():
        line = line.strip()
        if line:
            try:
                raw_items.append(json.loads(line))
            except Exception:
                pass

    # For true_false challenges, deduplicate by (statement, model_name) so repeated
    # generate runs don't inflate the count, but each model still gets its own row
    # per question.
    if challenge.task_type == "true_false":
        _seen: set[tuple] = set()
        items = []
        for _item in raw_items:
            _ref = _item.get("reference_answer", "")
            try:
                _ref_obj = json.loads(_ref) if isinstance(_ref, str) else _ref
                _stmt = _ref_obj.get("statement", "") if isinstance(_ref_obj, dict) else ""
            except Exception:
                _stmt = ""
            _mn = _item.get("model_name") or _item.get("model") or ""
            if _stmt:
                _key = (_stmt, _mn)
                if _key in _seen:
                    continue
                _seen.add(_key)
            items.append(_item)
    else:
        items = raw_items

    filtered_reasons = {"invalid_topic": 0, "unsupported_stance": 0, "no_replies": 0, "unsupported_language": 0}
    filtered_count = 0
    scored = []
    per_model_scores: dict[str, list[float]] = {}
    per_model_latencies: dict[str, list[float]] = {}
    per_model_error_counts: dict[str, int] = {}
    tested_models: list[str] = []
    judge_model_name = None
    selected_aspects = list(default_labels)
    leaderboard_target = "model"
    leaderboard_sample_size = 10
    leaderboard_sample_scope = "overall"
    model_api_base_url = None

    for item in items:
        model_name = str(item.get("model_name") or "")
        if model_name and model_name not in per_model_scores:
            per_model_scores[model_name] = []
            per_model_latencies[model_name] = []
            per_model_error_counts[model_name] = 0
            tested_models.append(model_name)
        if isinstance(item.get("score"), (int, float)):
            scored.append(float(item["score"]))
            if model_name:
                per_model_scores[model_name].append(float(item["score"]))
        if item.get("response_error") and model_name:
            per_model_error_counts[model_name] += 1
        if isinstance(item.get("latency_ms"), (int, float)) and model_name:
            per_model_latencies[model_name].append(float(item["latency_ms"]))
        if not judge_model_name and item.get("judge_model"):
            judge_model_name = item.get("judge_model")
        if item.get("selected_aspects") and isinstance(item.get("selected_aspects"), list):
            selected_aspects = item["selected_aspects"] or selected_aspects
        if item.get("leaderboard_target") in {"model", "participant"}:
            leaderboard_target = item["leaderboard_target"]
        if isinstance(item.get("leaderboard_sample_size"), int):
            leaderboard_sample_size = max(1, min(int(item["leaderboard_sample_size"]), 100))
        if item.get("leaderboard_sample_scope") in {"overall", "per_aspect"}:
            leaderboard_sample_scope = item["leaderboard_sample_scope"]
        if item.get("model_api_base_url"):
            model_api_base_url = item["model_api_base_url"]

        if item.get("task_type") != "stance_analysis":
            continue

        topic = (item.get("topic") or "").strip()
        stance = (item.get("stance") or "").strip().lower()
        thread_post = (item.get("thread_post") or topic or "").strip()
        thread_replies = (item.get("thread_replies") or item.get("content") or "").strip()

        reject_reason = None
        if _is_rejectable_stance_topic(topic, item.get("full_content", "") or thread_post):
            reject_reason = "invalid_topic"
        elif stance not in {"pro", "con", "支持", "反對"}:
            reject_reason = "unsupported_stance"
        elif not thread_replies:
            reject_reason = "no_replies"
        elif not _is_supported_stance_language(thread_post):
            reject_reason = "unsupported_language"

        if reject_reason:
            filtered_count += 1
            filtered_reasons[reject_reason] += 1

    avg_score = round(sum(scored) / len(scored), 3) if scored else None
    pass_flag = (avg_score >= 3.0) if avg_score is not None else None
    per_model = {
        model_name: {
            "avg_score": round(sum(scores) / len(scores), 3) if scores else None,
            "count": len(scores),
            "avg_latency_ms": round(sum(per_model_latencies[model_name]) / len(per_model_latencies[model_name]), 1)
            if per_model_latencies.get(model_name)
            else None,
            "error_count": per_model_error_counts.get(model_name, 0),
        }
        for model_name, scores in per_model_scores.items()
    }
    aspect_breakdown = _collect_aspect_breakdown(
        items, tested_models,
        labels=selected_aspects, use_keyword=False,
    )
    leaderboard = _sort_leaderboard_rows([
        {
            "entity_type": "model",
            "model_name": model_name_key,
            "display_name": model_name_key,
            **stats,
            "aspect_scores": aspect_breakdown.get(model_name_key, {}).get("scores", {}),
            "aspect_counts": aspect_breakdown.get(model_name_key, {}).get("counts", {}),
        }
        for model_name_key, stats in per_model.items()
    ])
    source_items = len({(item.get("task_type"), item.get("title"), item.get("topic"), item.get("content"), item.get("reference_answer")) for item in items})

    # Within-run dedup count (all rows from one generate share the same value)
    run_dedup_count = 0
    for it in items:
        v = it.get("_run_dedup_count")
        if isinstance(v, int) and v > 0:
            run_dedup_count = v
            break

    # Trim heavy fields from each item to reduce response size dramatically.
    # UI shows truncated content; full row available via /results/{index}.
    def _light(it: dict) -> dict:
        out = dict(it)
        # Drop fields used only in detail expand (lazy-loadable)
        out.pop("request_messages", None)
        out.pop("full_content", None)
        # Truncate long text fields
        c = out.get("content") or ""
        if len(c) > 300:
            out["content"] = c[:300] + "…"
        rt = out.get("response_text") or ""
        if len(rt) > 500:
            out["response_text"] = rt[:500] + "…"
        rsn = out.get("reasoning") or ""
        if len(rsn) > 300:
            out["reasoning"] = rsn[:300] + "…"
        return out

    items_light = [_light(it) for it in items]

    result = {
        "items": items_light,
        "total_items": len(items),
        "source_items": source_items,
        "filtered_items": filtered_count,
        "filtered_reasons": filtered_reasons,
        "run_dedup_count": run_dedup_count,
        "avg_score": avg_score,
        "pass": pass_flag,
        "aspect_labels": selected_aspects,
        "selected_aspects": selected_aspects,
        "leaderboard_target": leaderboard_target,
        "leaderboard_sample_size": leaderboard_sample_size,
        "leaderboard_sample_scope": leaderboard_sample_scope,
        "model_api_base_url": model_api_base_url,
        "tested_models": tested_models,
        "judge_model_name": judge_model_name,
        "per_model": per_model,
        "leaderboard": leaderboard,
    }

    # Cache 60s with fingerprint so re-generates invalidate automatically
    _results_cache[cache_key] = {
        "data": result,
        "fingerprint": fingerprint,
        "expires_at": time.monotonic() + 60,
    }
    return result


@router.get("/challenges/{uuid}/question-list")
async def get_question_list(uuid: str, db: AsyncSession = Depends(get_db)):
    """Return the deduplicated question list as JSON for use in selection UI."""
    challenge = await ChallengeCRUD.get_by_uuid(db, uuid)
    if not challenge:
        raise HTTPException(status_code=404, detail="Challenge not found")
    if not challenge.results_jsonl:
        return {"questions": []}

    items = []
    for line in challenge.results_jsonl.splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            items.append(json.loads(line))
        except Exception:
            continue

    seen_keys: set[tuple] = set()
    seen_statements: set[str] = set()
    unique_items = []
    for item in items:
        ref_value = item.get("reference_answer")
        if isinstance(ref_value, (dict, list)):
            ref_value = json.dumps(ref_value, ensure_ascii=False, sort_keys=True)
        statement_key: str | None = None
        if item.get("task_type") == "true_false":
            try:
                ref_obj = json.loads(ref_value) if isinstance(ref_value, str) else ref_value
                stmt = ref_obj.get("statement", "") if isinstance(ref_obj, dict) else ""
                if stmt and "{" not in stmt:
                    statement_key = stmt
            except Exception:
                pass
        if statement_key and statement_key in seen_statements:
            continue
        dedupe_key = (item.get("task_type"), item.get("title"), item.get("topic"), item.get("content"), ref_value)
        if dedupe_key in seen_keys:
            continue
        if statement_key:
            seen_statements.add(statement_key)
        seen_keys.add(dedupe_key)
        unique_items.append(item)

    # Build P-value map from model results (deduplicated per model)
    from collections import defaultdict
    _stmt_stats: dict[str, dict] = defaultdict(lambda: {"correct": 0, "total": 0})
    _model_rows: dict[str, list] = defaultdict(list)
    for _row in items:
        _mn = _row.get("model_name") or _row.get("model") or "__unknown__"
        _model_rows[_mn].append(_row)
    for _mn, _rows in _model_rows.items():
        _seen: set[str] = set()
        for _row in _rows:
            _s = ""
            _rr = _row.get("reference_answer", "")
            try:
                _ro = json.loads(_rr) if isinstance(_rr, str) else _rr
                if isinstance(_ro, dict):
                    _s = _ro.get("statement", "")
            except Exception:
                pass
            if not _s:
                _pv = _row.get("prompt", "")
                _s = (_pv.strip().splitlines()[-1].strip() if _pv else "")
            if not _s or _s in _seen:
                continue
            _seen.add(_s)
            _stmt_stats[_s]["total"] += 1
            if (_row.get("score") or 0) >= 4:
                _stmt_stats[_s]["correct"] += 1
    if challenge.participant_scores_jsonl:
        for _line in challenge.participant_scores_jsonl.splitlines():
            _line = _line.strip()
            if not _line:
                continue
            try:
                _sub = json.loads(_line)
            except Exception:
                continue
            _seen_sub: set[str] = set()
            for _pitem in _sub.get("items", []):
                _s = _pitem.get("statement", "").strip()
                if not _s or _s in _seen_sub:
                    continue
                _seen_sub.add(_s)
                _stmt_stats[_s]["total"] += 1
                if _pitem.get("correct"):
                    _stmt_stats[_s]["correct"] += 1
    _pval_map: dict[str, float | None] = {
        s: round(info["correct"] / info["total"], 3) if info["total"] > 0 else None
        for s, info in _stmt_stats.items()
    }

    def _pcat(p: float | None) -> str:
        if p is None: return "unknown"
        if p > 0.7: return "easy"
        if p < 0.4: return "hard"
        return "ideal"

    questions = []
    for idx, item in enumerate(unique_items):
        task_type = item.get("task_type", "qa")
        statement = ""
        ref_raw = item.get("reference_answer", "")
        try:
            ref_obj = json.loads(ref_raw) if isinstance(ref_raw, str) else ref_raw
            if isinstance(ref_obj, dict):
                statement = ref_obj.get("statement", "") or ref_obj.get("question", "")
        except Exception:
            pass
        if not statement:
            prompt = item.get("prompt", "")
            statement = prompt.strip().splitlines()[-1].strip() if prompt else ""
        if not statement or "{" in statement:
            continue
        p = _pval_map.get(statement)
        title = item.get("title", "") or item.get("topic", "")
        aspect = item.get("aspect", "") or _infer_aspect_from_text(title)
        questions.append({
            "index": idx,
            "task_type": task_type,
            "title": title,
            "aspect": aspect,
            "statement": statement,
            "p_value": p,
            "p_category": _pcat(p),
        })

    return {"questions": questions, "total": len(questions)}


@router.get("/challenges/{uuid}/export-questions")
async def export_questions(
    uuid: str,
    sample_size: Optional[int] = Query(None, ge=1, le=100),
    categories: Optional[str] = Query(None),  # comma-separated: ideal,easy,hard
    indices: Optional[str] = Query(None),      # comma-separated 0-based indices from question-list
    db: AsyncSession = Depends(get_db),
):
    """Export questions/tasks for human answering."""
    from io import BytesIO

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation
    except ModuleNotFoundError as exc:
        raise HTTPException(
            status_code=500,
            detail="缺少 openpyxl，無法匯出 XLSX 題目。",
        ) from exc

    challenge = await ChallengeCRUD.get_by_uuid(db, uuid)
    if not challenge:
        raise HTTPException(status_code=404, detail="Challenge not found")
    if not challenge.results_jsonl:
        raise HTTPException(status_code=400, detail="No generated questions. Run generation first.")

    items = []
    for line in challenge.results_jsonl.splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            items.append(json.loads(line))
        except Exception:
            continue

    if not items:
        raise HTTPException(status_code=400, detail="No items found.")

    unique_items = []
    seen_keys: set[tuple] = set()
    seen_statements: set[str] = set()
    for item in items:
        ref_value = item.get("reference_answer")
        if isinstance(ref_value, (dict, list)):
            ref_value = json.dumps(ref_value, ensure_ascii=False, sort_keys=True)
        # For true_false, also deduplicate by statement text
        statement_key: str | None = None
        if item.get("task_type") == "true_false":
            try:
                ref_obj = json.loads(ref_value) if isinstance(ref_value, str) else ref_value
                stmt = ref_obj.get("statement", "") if isinstance(ref_obj, dict) else ""
                if stmt:
                    statement_key = stmt
            except Exception:
                pass
        if statement_key and statement_key in seen_statements:
            continue
        dedupe_key = (
            item.get("task_type"),
            item.get("title"),
            item.get("topic"),
            item.get("content"),
            ref_value,
        )
        if dedupe_key in seen_keys:
            continue
        if statement_key:
            seen_statements.add(statement_key)
        seen_keys.add(dedupe_key)
        unique_items.append(item)

    if not unique_items:
        raise HTTPException(status_code=400, detail="No unique questions found.")

    # Filter by manually selected indices
    if indices:
        try:
            idx_set = {int(i.strip()) for i in indices.split(",") if i.strip().lstrip("-").isdigit()}
            selected = [it for i, it in enumerate(unique_items) if i in idx_set]
            # preserve the order of indices as specified
            idx_order = [int(i.strip()) for i in indices.split(",") if i.strip().lstrip("-").isdigit()]
            idx_map = {i: it for i, it in enumerate(unique_items) if i in idx_set}
            unique_items = [idx_map[i] for i in idx_order if i in idx_map]
        except Exception:
            pass
        if not unique_items:
            raise HTTPException(status_code=400, detail="指定的題目編號無效。")

    # Filter by P-value categories when requested
    if categories:
        from collections import defaultdict
        _selected = {c.strip().lower() for c in categories.split(",") if c.strip()}
        _stmt_stats: dict[str, dict] = defaultdict(lambda: {"correct": 0, "total": 0})
        _model_rows: dict[str, list] = defaultdict(list)
        for _row in items:
            _mn = _row.get("model_name") or _row.get("model") or ""
            if _mn:
                _model_rows[_mn].append(_row)
        for _mn, _rows in _model_rows.items():
            _seen: set[str] = set()
            for _row in _rows:
                _stmt = ""
                _ref_raw = _row.get("reference_answer", "")
                try:
                    _ref_obj = json.loads(_ref_raw) if isinstance(_ref_raw, str) else _ref_raw
                    if isinstance(_ref_obj, dict):
                        _stmt = _ref_obj.get("statement", "")
                except Exception:
                    pass
                if not _stmt:
                    _pv = _row.get("prompt", "")
                    _stmt = (_pv.strip().splitlines()[-1].strip() if _pv else "")
                if not _stmt or _stmt in _seen:
                    continue
                _seen.add(_stmt)
                _stmt_stats[_stmt]["total"] += 1
                if (_row.get("score") or 0) >= 4:
                    _stmt_stats[_stmt]["correct"] += 1
        if challenge.participant_scores_jsonl:
            for _line in challenge.participant_scores_jsonl.splitlines():
                _line = _line.strip()
                if not _line:
                    continue
                try:
                    _sub = json.loads(_line)
                except Exception:
                    continue
                for _pitem in _sub.get("items", []):
                    _stmt = _pitem.get("statement", "").strip()
                    if not _stmt:
                        continue
                    _stmt_stats[_stmt]["total"] += 1
                    if _pitem.get("correct"):
                        _stmt_stats[_stmt]["correct"] += 1

        def _pcat(p: Optional[float]) -> str:
            if p is None:
                return "unknown"
            if p > 0.7:
                return "easy"
            if p < 0.4:
                return "hard"
            return "ideal"

        _pval_map: dict[str, float] = {
            s: (info["correct"] / info["total"]) for s, info in _stmt_stats.items() if info["total"] > 0
        }

        def _item_stmt(it: dict) -> str:
            _s = ""
            _rr = it.get("reference_answer", "")
            try:
                _ro = json.loads(_rr) if isinstance(_rr, str) else _rr
                if isinstance(_ro, dict):
                    _s = _ro.get("statement", "")
            except Exception:
                pass
            if not _s:
                _pv = it.get("prompt", "")
                _s = (_pv.strip().splitlines()[-1].strip() if _pv else "")
            return _s

        unique_items = [
            it for it in unique_items
            if _pcat(_pval_map.get(_item_stmt(it))) in _selected
        ]
        if not unique_items:
            raise HTTPException(status_code=400, detail=f"篩選後無符合類別 {categories} 的題目。")

    effective_sample_size = sample_size
    if effective_sample_size is None:
        first_size = unique_items[0].get("leaderboard_sample_size")
        if isinstance(first_size, int):
            effective_sample_size = max(1, min(first_size, 100))
    if effective_sample_size is not None:
        unique_items = _diversify_by_article(unique_items, effective_sample_size)

    task_type = unique_items[0].get("task_type", "qa")

    if task_type == "stance_analysis":
        wb = Workbook()
        ws = wb.active
        ws.title = "Questions"

        headers = ["#", "立場（支持/反對）", "原文文章", "別人留言回覆"]
        ws.append(headers)

        header_fill = PatternFill(fill_type="solid", fgColor="E8EEF9")
        body_font = Font(name="Microsoft JhengHei", size=11)
        header_font = Font(name="Microsoft JhengHei", size=11, bold=True)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        stance_validation = DataValidation(
            type="list",
            formula1='"支持,反對"',
            allow_blank=True,
        )
        ws.add_data_validation(stance_validation)

        for i, item in enumerate(unique_items, 1):
            thread_post = item.get("thread_post", "") or item.get("topic", "")
            thread_replies = item.get("thread_replies", "") or item.get("content", "")
            ws.append([i, "", thread_post, thread_replies])
            stance_validation.add(ws.cell(row=i + 1, column=2))
            for col in range(1, 5):
                ws.cell(row=i + 1, column=col).font = body_font
            ws.cell(row=i + 1, column=1).alignment = Alignment(horizontal="center", vertical="top")
            ws.cell(row=i + 1, column=2).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=i + 1, column=3).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=i + 1, column=4).alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[i + 1].height = 72

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.sheet_view.showGridLines = True
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 24
        ws.column_dimensions["D"].width = 22
        ws.row_dimensions[1].height = 28

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = _safe_download_filename(challenge.name, "_questions.xlsx")
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": _build_content_disposition(filename)},
        )

    if task_type == "true_false":
        wb = Workbook()
        ws = wb.active
        ws.title = "Questions"
        ws.append(["#", "陳述句", "TRUE / FALSE"])

        header_fill = PatternFill(fill_type="solid", fgColor="E8EEF9")
        body_font = Font(name="Microsoft JhengHei", size=11)
        header_font = Font(name="Microsoft JhengHei", size=11, bold=True)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        tf_validation = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
        ws.add_data_validation(tf_validation)

        seq = 0
        for item in unique_items:
            ref_raw = item.get("reference_answer", "")
            statement = ""
            try:
                ref = json.loads(ref_raw) if isinstance(ref_raw, str) else ref_raw
                if isinstance(ref, dict):
                    statement = ref.get("statement", "")
            except Exception:
                pass
            if not statement:
                prompt = item.get("prompt", "")
                if prompt:
                    statement = prompt.strip().splitlines()[-1].strip()

            if not statement or "{" in statement:
                continue

            seq += 1
            ws.append([seq, statement, ""])
            row = seq + 1
            for col in range(1, 4):
                ws.cell(row=row, column=col).font = body_font
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="center")
            answer_cell = ws.cell(row=row, column=3)
            answer_cell.alignment = Alignment(horizontal="center", vertical="center")
            tf_validation.add(answer_cell)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.sheet_view.showGridLines = True
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 14
        ws.row_dimensions[1].height = 28

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = _safe_download_filename(challenge.name, "_questions.xlsx")
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": _build_content_disposition(filename)},
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Questions"
    ws.append(["#", "文章標題", "問題", "回答"])

    header_fill = PatternFill(fill_type="solid", fgColor="E8EEF9")
    body_font = Font(name="Microsoft JhengHei", size=11)
    header_font = Font(name="Microsoft JhengHei", size=11, bold=True)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, item in enumerate(unique_items, 1):
        ref_raw = item.get("reference_answer", "")
        question = ""
        try:
            ref = json.loads(ref_raw) if isinstance(ref_raw, str) else ref_raw
            if isinstance(ref, dict):
                stem = ref.get("question", "")
                if ref.get("type") == "選擇題" and ref.get("choices"):
                    choices_text = "\n".join(
                        f"{k}. {v}" for k, v in sorted(ref["choices"].items()) if v
                    )
                    question = f"{stem}\n{choices_text}" if stem else choices_text
                else:
                    question = stem
        except Exception:
            pass

        ws.append([i, item.get("title", ""), question, ""])
        for col in range(1, 5):
            ws.cell(row=i + 1, column=col).font = body_font
        ws.cell(row=i + 1, column=1).alignment = Alignment(horizontal="center", vertical="top")
        ws.cell(row=i + 1, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=i + 1, column=3).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=i + 1, column=4).alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = True
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 56
    ws.column_dimensions["D"].width = 24
    ws.row_dimensions[1].height = 28

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = _safe_download_filename(challenge.name, "_questions.xlsx")
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": _build_content_disposition(filename)},
    )

@router.post("/challenges/{uuid}/score-answers")
async def score_human_answers(
    uuid: str,
    file: UploadFile = File(...),
    model_name: str = Form("gpt-oss-20b"),
    db: AsyncSession = Depends(get_db),
):
    """Score human answers uploaded via CSV/XLSX against stored reference answers."""
    import csv
    import os
    from io import BytesIO
    from pathlib import Path
    from ...qual.prompts.judge import JUDGE_SYSTEM_PROMPT, JUDGE_SCORE_PROMPT
    from openai import AsyncOpenAI, RateLimitError
    from datetime import datetime, timezone

    challenge = await ChallengeCRUD.get_by_uuid(db, uuid)
    if not challenge:
        raise HTTPException(status_code=404, detail="Challenge not found")
    if not challenge.results_jsonl:
        raise HTTPException(status_code=400, detail="No reference questions found.")

    ref_rows = []
    for line in challenge.results_jsonl.splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            ref_rows.append(json.loads(line))
        except Exception:
            pass

    content = await file.read()
    filename = (file.filename or "").lower()

    if filename.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError as exc:
            raise HTTPException(
                status_code=500,
                detail="缺少 openpyxl，無法讀取 XLSX 受試者答案。",
            ) from exc
        wb = load_workbook(BytesIO(content), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(status_code=400, detail="XLSX is empty.")
        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        answer_rows = []
        for row in rows[1:]:
            row_values = list(row)
            if not any(v not in (None, "") for v in row_values):
                continue
            answer_rows.append({
                headers[i]: "" if i >= len(row_values) or row_values[i] is None else str(row_values[i]).strip()
                for i in range(len(headers))
            })
    else:
        try:
            text = content.decode("utf-8-sig")
        except Exception:
            text = content.decode("utf-8", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        answer_rows = list(reader)

    if not answer_rows:
        raise HTTPException(status_code=400, detail="Answer file is empty or invalid.")

    litellm_url = os.getenv("LITELLM_URL", "https://llmgw.elandai.cloud")
    api_key = os.getenv("LITELLM_API_KEY", "")
    client = AsyncOpenAI(base_url=f"{litellm_url}/v1", api_key=api_key or "no-key")
    semaphore = asyncio.Semaphore(4)

    task_type = ref_rows[0].get("task_type", "qa") if ref_rows else "qa"

    def _extract_rate_limit_delay(error_text: str) -> float:
        marker = "Limit resets at:"
        if marker in error_text:
            reset_text = error_text.split(marker, 1)[1].strip().rstrip("'").rstrip("}")
            if reset_text.endswith("UTC"):
                reset_text = reset_text[:-3].strip()
            try:
                reset_at = datetime.strptime(reset_text, "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)
                delay = (reset_at - datetime.now(timezone.utc)).total_seconds() + 1.0
                return max(1.0, min(delay, 30.0))
            except ValueError:
                pass
        return 5.0

    async def call_judge(task_type_str: str, prompt_text: str, human_answer: str, reference_answer: str, rubric: str) -> dict:
        prompt = JUDGE_SCORE_PROMPT.format(
            task_type=task_type_str,
            prompt=prompt_text,
            reference_answer=reference_answer,
            response=human_answer,
            scoring_rubric=rubric,
        )
        async with semaphore:
            for attempt in range(3):
                try:
                    resp = await client.chat.completions.create(
                        model=model_name,
                        messages=[
                            {"role": "system", "content": JUDGE_SYSTEM_PROMPT},
                            {"role": "user", "content": prompt},
                        ],
                        temperature=0.0,
                    )
                    text = (resp.choices[0].message.content or "").strip()
                    if text.startswith("```"):
                        lines = text.splitlines()
                        lines = lines[1:] if lines[0].startswith("```") else lines
                        lines = lines[:-1] if lines and lines[-1].strip() == "```" else lines
                        text = "\n".join(lines).strip()
                    return json.loads(text)
                except RateLimitError as e:
                    if attempt >= 2:
                        return {"score": None, "reasoning": f"評分失敗：{e}"}
                    await asyncio.sleep(_extract_rate_limit_delay(str(e)))
                except Exception as e:
                    return {"score": None, "reasoning": f"評分失敗：{e}"}

    participant_name = Path(file.filename or "受試者").stem or "受試者"
    results = []
    tasks = []
    matched_meta = []

    # For true_false: build a statement → ref lookup for content-based matching
    tf_statement_index: dict[str, dict] = {}
    if task_type == "true_false":
        for r in ref_rows:
            try:
                ref_obj = json.loads(r.get("reference_answer", "")) if isinstance(r.get("reference_answer"), str) else r.get("reference_answer", {})
                stmt = ref_obj.get("statement", "") if isinstance(ref_obj, dict) else ""
            except Exception:
                stmt = ""
            if not stmt:
                prompt_val = r.get("prompt", "")
                stmt = prompt_val.strip().splitlines()[-1].strip() if prompt_val else ""
            if stmt:
                tf_statement_index[stmt] = r

    for ans_row in answer_rows:
        if task_type == "true_false":
            ans_statement = (ans_row.get("陳述句", "") or "").strip()
            ref = tf_statement_index.get(ans_statement)
            if not ref:
                continue
        else:
            idx_str = (ans_row.get("#", "") or "").strip()
            try:
                idx = int(float(idx_str)) - 1
            except Exception:
                continue
            if idx < 0 or idx >= len(ref_rows):
                continue
            ref = ref_rows[idx]

        rubric = ref.get("scoring_rubric", "")

        if task_type == "stance_analysis":
            human_stance = ((ans_row.get("立場（支持/反對）", "") or ans_row.get("立場（支持/反對/中立/混合）", "")) or "").strip()
            human_evidence = (ans_row.get("依據", "") or "").strip()
            if not human_stance and not human_evidence:
                continue
            human_answer_str = f"立場：{human_stance}\n依據：{human_evidence}"
            topic = (ans_row.get("主題描述", "") or "").strip() or ref.get("topic", "")
            prompt_text = f"請分析以下討論串主題作者的立場。\n主題描述：{topic}\n討論串內容：{ref.get('content', '')}"
            ref_raw = ref.get("reference_answer", "")
            ref_answer_str = ref_raw
            try:
                ref_obj = json.loads(ref_raw) if isinstance(ref_raw, str) else ref_raw
                if isinstance(ref_obj, dict):
                    ref_answer_str = f"立場：{ref_obj.get('stance', '')}\n依據：{ref_obj.get('evidence', '')}"
            except Exception:
                pass
            display_question = thread_post
        elif task_type == "true_false":
            human_answer_str = (ans_row.get("TRUE / FALSE", "") or "").strip().upper()
            if not human_answer_str:
                continue
            ref_raw = ref.get("reference_answer", "")
            ref_answer_str = ""
            statement = ""
            try:
                ref_obj = json.loads(ref_raw) if isinstance(ref_raw, str) else ref_raw
                if isinstance(ref_obj, dict):
                    ref_answer_str = ref_obj.get("answer", "").upper()
                    statement = ref_obj.get("statement", "")
            except Exception:
                ref_answer_str = ref_raw
            if not statement:
                prompt_val = ref.get("prompt", "")
                statement = prompt_val.strip().splitlines()[-1].strip() if prompt_val else ""
            prompt_text = statement
            display_question = statement
        else:
            human_answer_str = (ans_row.get("回答", "") or "").strip()
            if not human_answer_str:
                continue
            question = (ans_row.get("問題", "") or "").strip()
            ref_raw = ref.get("reference_answer", "")
            ref_answer_str = ""
            try:
                ref_obj = json.loads(ref_raw) if isinstance(ref_raw, str) else ref_raw
                ref_answer_str = ref_obj.get("answer", "") if isinstance(ref_obj, dict) else str(ref_obj)
                if not question:
                    question = ref_obj.get("question", "") if isinstance(ref_obj, dict) else ""
            except Exception:
                ref_answer_str = ref_raw
            prompt_text = f"問題：{question}"
            display_question = question

        matched_meta.append((ref, display_question, ref_answer_str, human_answer_str))
        if task_type == "true_false":
            tasks.append(None)
        else:
            tasks.append(call_judge(task_type, prompt_text, human_answer_str, ref_answer_str, rubric))

    judge_tasks = [t for t in tasks if t is not None]
    judge_results = await asyncio.gather(*judge_tasks)
    judge_iter = iter(judge_results)

    for (ref, display_question, ref_answer_str, human_answer_str), task in zip(matched_meta, tasks):
        if task_type == "true_false":
            correct = human_answer_str.upper() == ref_answer_str.upper()
            score_result = {"score": 5 if correct else 1, "reasoning": "正確" if correct else f"答案應為 {ref_answer_str}"}
        else:
            score_result = next(judge_iter)
        results.append({
            "task_type": task_type,
            "title": ref.get("title", "") or ref.get("topic", ""),
            "question": display_question,
            "human_answer": human_answer_str,
            "reference_answer": ref_answer_str,
            "score": score_result.get("score"),
            "reasoning": score_result.get("reasoning", ""),
        })

    total = len(results)
    scores_only = [r["score"] for r in results if r["score"] is not None]
    avg = round(sum(scores_only) / len(scores_only), 2) if scores_only else None

    # Persist this participant's scored items so P-values can be computed later
    from datetime import datetime, timezone as _tz
    score_record = json.dumps({
        "participant": participant_name,
        "scored_at": datetime.now(_tz.utc).isoformat(),
        "items": [
            {
                "statement": r.get("question", ""),
                "answer": r.get("human_answer", ""),
                "ref_answer": r.get("reference_answer", ""),
                "correct": r.get("score") == 5,
                "title": r.get("title", ""),
            }
            for r in results
        ],
    }, ensure_ascii=False)
    existing = challenge.participant_scores_jsonl or ""
    challenge.participant_scores_jsonl = (existing + "\n" + score_record).lstrip("\n")
    await db.commit()

    participant_aspect_items = [
        {
            "model_name": participant_name,
            "aspect": _infer_aspect_from_text(
                str(row.get("title") or ""),
                str(row.get("prompt") or ""),
                str(row.get("reference_answer") or ""),
            ),
            "score": row.get("score"),
        }
        for row in results
    ]
    aspect_breakdown = _collect_aspect_breakdown(
        participant_aspect_items, [participant_name],
        labels=_DEFAULT_ASPECT_LABELS, use_keyword=False,
    )
    leaderboard_rows = _sort_leaderboard_rows([{
        "entity_type": "participant",
        "model_name": participant_name,
        "display_name": participant_name,
        "avg_score": avg,
        "count": total,
        "avg_latency_ms": None,
        "error_count": sum(1 for row in results if row.get("score") is None),
        "aspect_scores": aspect_breakdown.get(participant_name, {}).get("scores", {}),
        "aspect_counts": aspect_breakdown.get(participant_name, {}).get("counts", {}),
    }])
    return {
        "participant_name": participant_name,
        "total": total,
        "avg_score": avg,
        "leaderboard_rows": leaderboard_rows,
        "items": results,
    }

@router.get("/challenges/{uuid}/p-values")
async def get_challenge_p_values(uuid: str, db: AsyncSession = Depends(get_db)):
    """Return per-question P-values for ALL generated questions; untested questions get p_value=null."""
    from collections import defaultdict

    challenge = await ChallengeCRUD.get_by_uuid(db, uuid)
    if not challenge:
        raise HTTPException(status_code=404, detail="Challenge not found")
    if not challenge.results_jsonl:
        return {"participants": [], "question_count": 0, "questions": []}

    # --- Step 1: collect ALL unique questions (same logic as question-list) ---
    all_rows: list[dict] = []
    for line in challenge.results_jsonl.splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            all_rows.append(json.loads(line))
        except Exception:
            continue

    seen_stmts_global: set[str] = set()
    all_questions: list[dict] = []  # {statement, title}
    for row in all_rows:
        ref_raw = row.get("reference_answer", "")
        try:
            ref_obj = json.loads(ref_raw) if isinstance(ref_raw, str) else ref_raw
            stmt = ref_obj.get("statement", "") if isinstance(ref_obj, dict) else ""
        except Exception:
            stmt = ""
        if not stmt or "{" in stmt:
            continue
        if stmt in seen_stmts_global:
            continue
        seen_stmts_global.add(stmt)
        all_questions.append({"statement": stmt, "title": row.get("title", "")})

    # --- Step 2: compute P-values from scored rows ---
    stmt_stats: dict[str, dict] = defaultdict(lambda: {"correct": 0, "total": 0})
    participants: list[str] = []

    model_rows: dict[str, list[dict]] = defaultdict(list)
    for row in all_rows:
        model_name = row.get("model_name") or row.get("model") or ""
        if model_name:
            model_rows[model_name].append(row)

    for model_name, rows in model_rows.items():
        participants.append(model_name)
        seen_per_model: set[str] = set()
        for row in rows:
            score = row.get("score")
            if score is None:
                continue  # not yet evaluated — skip for P-value computation
            ref_raw = row.get("reference_answer", "")
            try:
                ref_obj = json.loads(ref_raw) if isinstance(ref_raw, str) else ref_raw
                stmt = ref_obj.get("statement", "") if isinstance(ref_obj, dict) else ""
            except Exception:
                stmt = ""
            if not stmt or stmt in seen_per_model:
                continue
            seen_per_model.add(stmt)
            stmt_stats[stmt]["total"] += 1
            if (score or 0) >= 4:
                stmt_stats[stmt]["correct"] += 1

    # --- Source 2: uploaded participant answer files ---
    if challenge.participant_scores_jsonl:
        for line in challenge.participant_scores_jsonl.splitlines():
            line = line.strip()
            if not line:
                continue
            try:
                sub = json.loads(line)
            except Exception:
                continue
            participants.append(sub.get("participant", "受試者"))
            for item in sub.get("items", []):
                stmt = item.get("statement", "").strip()
                if not stmt:
                    continue
                stmt_stats[stmt]["total"] += 1
                if item.get("correct"):
                    stmt_stats[stmt]["correct"] += 1

    # --- Step 3: merge all_questions with P-value stats ---
    questions = sorted(
        [
            {
                "statement": q["statement"],
                "title": q["title"],
                "correct": stmt_stats[q["statement"]]["correct"],
                "total": stmt_stats[q["statement"]]["total"],
                "p_value": (
                    round(stmt_stats[q["statement"]]["correct"] / stmt_stats[q["statement"]]["total"], 3)
                    if stmt_stats[q["statement"]]["total"] > 0 else None
                ),
            }
            for q in all_questions
        ],
        key=lambda q: (q["p_value"] is None, q["p_value"] or 0),
    )

    return {
        "participants": participants,
        "question_count": len(questions),
        "questions": questions,
    }


@router.delete("/challenges/{uuid}")
async def delete_challenge(uuid: str, db: AsyncSession = Depends(get_db)):
    """Delete a challenge."""
    challenge = await ChallengeCRUD.get_by_uuid(db, uuid)
    if not challenge:
        raise HTTPException(status_code=404, detail="Challenge not found")
    await ChallengeCRUD.delete(db, uuid)
    await db.commit()
    return {"message": "Challenge deleted", "uuid": uuid}


@router.get("/challenges/{uuid}/export/google-form-script")
async def export_google_form_script(
    uuid: str,
    sample_size: int = Query(20, ge=1, le=100),
    indices: Optional[str] = Query(None),  # comma-separated 0-based indices; overrides sample_size
    db: AsyncSession = Depends(get_db),
):
    """Export true_false questions as a Google Apps Script that creates a Google Form quiz."""
    challenge = await ChallengeCRUD.get_by_uuid(db, uuid)
    if not challenge:
        raise HTTPException(status_code=404, detail="Challenge not found")
    if challenge.task_type != "true_false":
        raise HTTPException(status_code=400, detail="只支援 true_false 題型匯出 Google 表單。")
    if not challenge.results_jsonl:
        raise HTTPException(status_code=400, detail="尚未生成題目，請先執行 Generate。")

    items = []
    for line in challenge.results_jsonl.splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            items.append(json.loads(line))
        except Exception:
            continue

    # Deduplicate by statement
    seen: set[str] = set()
    unique: list[dict] = []
    for item in items:
        ref_raw = item.get("reference_answer", "")
        try:
            ref = json.loads(ref_raw) if isinstance(ref_raw, str) else ref_raw
            statement = ref.get("statement", "") if isinstance(ref, dict) else ""
        except Exception:
            statement = ""
        if not statement or statement in seen:
            continue
        seen.add(statement)
        unique.append(item)

    if not unique:
        raise HTTPException(status_code=400, detail="找不到有效的是非題題目。")

    if indices:
        try:
            idx_order = [int(i.strip()) for i in indices.split(",") if i.strip().lstrip("-").isdigit()]
            idx_map = {i: it for i, it in enumerate(unique)}
            unique = [idx_map[i] for i in idx_order if i in idx_map]
        except Exception:
            pass
        if not unique:
            raise HTTPException(status_code=400, detail="指定的題目編號無效。")
    else:
        unique = _diversify_by_article(unique, sample_size)

    # Build question list for the script
    questions = []
    for item in unique:
        ref_raw = item.get("reference_answer", "")
        try:
            ref = json.loads(ref_raw) if isinstance(ref_raw, str) else ref_raw
        except Exception:
            continue
        statement = ref.get("statement", "")
        answer = ref.get("answer", "").upper()
        if not statement or answer not in ("TRUE", "FALSE"):
            continue
        explanation = ref.get("explanation", "")
        questions.append({"statement": statement, "answer": answer, "explanation": explanation})

    if not questions:
        raise HTTPException(status_code=400, detail="無法解析題目內容。")

    form_title = challenge.name or "台灣知識是非題測驗"

    # Render questions as JS array literal
    q_lines = []
    for q in questions:
        stmt = q["statement"].replace("\\", "\\\\").replace('"', '\\"')
        expl = (q.get("explanation") or "").replace("\\", "\\\\").replace('"', '\\"')
        q_lines.append(f'  {{statement: "{stmt}", answer: "{q["answer"]}", explanation: "{expl}"}}')
    questions_js = "[\n" + ",\n".join(q_lines) + "\n]"

    script = f"""\
// 使用方式：
// 1. 前往 https://script.google.com/ 建立新專案
// 2. 刪除預設內容，貼上此程式碼並儲存（Ctrl+S）
// 3. 上方函式選單選擇「main」，點選「執行」
// 4. 首次執行需授權（需勾選 Google Forms 權限），授權後表單自動建立
// 5. 完成後在「執行記錄」中可看到表單填寫/編輯網址
//
// ⚠️ 注意：請勿點「部署」，直接點「執行」即可

/**
 * @OnlyCurrentDoc
 */

function main() {{
  createQuizForm();
}}

function createQuizForm() {{
  var title = "{form_title.replace('"', '\\"')}";
  var form = FormApp.create(title);
  form.setTitle(title);
  form.setDescription("請判斷以下關於台灣的陳述是否正確。回答 TRUE（正確）或 FALSE（錯誤）。");
  form.setIsQuiz(true);
  form.setCollectEmail(false);
  form.setLimitOneResponsePerUser(false);  // 允許任何人用連結填答，不需登入
  form.setShowLinkToRespondAgain(true);

  var questions = {questions_js};

  questions.forEach(function(q, i) {{
    var item = form.addMultipleChoiceItem();
    item.setTitle((i + 1) + ". " + q.statement);
    item.setRequired(true);
    var trueChoice  = item.createChoice("TRUE（正確）",  q.answer === "TRUE");
    var falseChoice = item.createChoice("FALSE（錯誤）", q.answer === "FALSE");
    item.setChoices([trueChoice, falseChoice]);
    item.setPoints(1);
    if (q.explanation) {{
      var fb = FormApp.createFeedback().setText(q.explanation).build();
      item.setFeedbackForCorrect(fb);
      item.setFeedbackForIncorrect(fb);
    }}
  }});

  // 開啟成績顯示：透過 Forms REST API 設定「立即發布成績」
  var token = ScriptApp.getOAuthToken();
  var apiUrl = "https://forms.googleapis.com/v1/forms/" + form.getId() + ":batchUpdate";
  var payload = JSON.stringify({{
    "requests": [{{
      "updateSettings": {{
        "settings": {{
          "quizSettings": {{ "isQuiz": true }}
        }},
        "updateMask": "quizSettings"
      }}
    }}]
  }});
  var response = UrlFetchApp.fetch(apiUrl, {{
    method: "post",
    contentType: "application/json",
    headers: {{ "Authorization": "Bearer " + token }},
    payload: payload,
    muteHttpExceptions: true
  }});

  // 開放「知道連結的人」均可存取（不需要向擁有者要求權限）
  DriveApp.getFileById(form.getId()).setSharing(
    DriveApp.Access.ANYONE_WITH_LINK,
    DriveApp.Permission.VIEW
  );

  Logger.log("表單網址（填寫）：" + form.getPublishedUrl());
  Logger.log("表單網址（編輯）：" + form.getEditUrl());
  Logger.log("");
  Logger.log("⚠️  若受試者看不到成績，請手動開啟：");
  Logger.log("   1. 點開表單編輯頁 → 右上角齒輪「設定」");
  Logger.log("   2. 點選「問卷」分頁");
  Logger.log("   3. 發布成績 → 選「立即（提交後立即顯示）」");
  Logger.log("   4. 儲存");
}}
"""

    filename = _safe_download_filename(challenge.name, "_google_form.gs")
    return StreamingResponse(
        iter([script]),
        media_type="text/plain; charset=utf-8",
        headers={"Content-Disposition": _build_content_disposition(filename)},
    )


@router.get("/challenges/{uuid}/download")
async def download_challenge(uuid: str, db: AsyncSession = Depends(get_db)):
    """Download challenge data as JSONL file."""
    challenge = await ChallengeCRUD.get_by_uuid(db, uuid)
    if not challenge:
        raise HTTPException(status_code=404, detail="Challenge not found")
    if not challenge.data_jsonl:
        raise HTTPException(status_code=404, detail="No data available")

    filename = f"{challenge.name.replace(' ', '_')}.jsonl"
    return StreamingResponse(
        iter([challenge.data_jsonl]),
        media_type="application/x-ndjson",
        headers={"Content-Disposition": _build_content_disposition(filename)},
    )












